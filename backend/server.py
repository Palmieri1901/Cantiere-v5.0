from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import io
import zipfile
import bcrypt
import jwt
from datetime import timedelta
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, date
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Cantiere Nautico API")

TOTAL_POSTI = 200


# ---------- AUTH (setup precoce per protezione router) ----------

from fastapi import Request, Depends, Response

JWT_ALGORITHM = "HS256"

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=8),
        "type": "access",
    }
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Non autenticato")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Token non valido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sessione scaduta")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token non valido")


api_router = APIRouter(prefix="/api")


# ---------- MODELS ----------

class Tariffe(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: "default")
    copertura_per_metro: float = 45.0
    # Alaggio/Varo a scaglioni per lunghezza
    alaggio_fino_5m: float = 90.0
    alaggio_oltre_5m_per_metro: float = 25.0
    varo_fino_5m: float = 90.0
    varo_oltre_5m_per_metro: float = 25.0
    antivegetativa_per_metro: float = 60.0
    # Tariffa sosta temporanea (a giorno)
    sosta_temporanea_giornaliera: float = 25.0
    # Manodopera motore a scaglioni di potenza HP
    motore_labor_2_15hp: float = 90.0
    motore_labor_fino_40hp: float = 180.0
    motore_labor_40_150hp: float = 320.0
    motore_labor_oltre_150hp: float = 550.0
    # Ricambi motore (costo unitario)
    costo_girante: float = 45.0
    costo_olio_motore: float = 12.0  # € al litro
    costo_filtro_olio: float = 18.0
    costo_candela: float = 12.0
    costo_termostato: float = 35.0
    costo_olio_piede: float = 25.0
    costo_anodi_interni: float = 40.0
    costo_anodi_esterni: float = 60.0
    costo_ingrassaggio: float = 30.0
    # Sosta
    sosta_dentro_per_metro: float = 180.0
    sosta_fuori_per_metro: float = 120.0
    # Sosta fuori sede: nessun costo sosta, ma movimentazione + taccaggio
    costo_movimentazione_per_metro: float = 25.0
    costo_taccaggio_per_metro: float = 20.0
    # Lavaggi ed extra
    costo_lavaggio_inizio_stagione: float = 80.0
    costo_lavaggio_fine_stagione: float = 80.0
    maggiorazione_scafo_sporco_per_metro: float = 15.0  # applicata se antivegetativa disattivata
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class TariffeUpdate(BaseModel):
    copertura_per_metro: Optional[float] = None
    alaggio_fino_5m: Optional[float] = None
    alaggio_oltre_5m_per_metro: Optional[float] = None
    varo_fino_5m: Optional[float] = None
    varo_oltre_5m_per_metro: Optional[float] = None
    antivegetativa_per_metro: Optional[float] = None
    motore_labor_2_15hp: Optional[float] = None
    motore_labor_fino_40hp: Optional[float] = None
    motore_labor_40_150hp: Optional[float] = None
    motore_labor_oltre_150hp: Optional[float] = None
    costo_girante: Optional[float] = None
    costo_olio_motore: Optional[float] = None
    costo_filtro_olio: Optional[float] = None
    costo_candela: Optional[float] = None
    costo_termostato: Optional[float] = None
    costo_olio_piede: Optional[float] = None
    costo_anodi_interni: Optional[float] = None
    costo_anodi_esterni: Optional[float] = None
    costo_ingrassaggio: Optional[float] = None
    sosta_dentro_per_metro: Optional[float] = None
    sosta_fuori_per_metro: Optional[float] = None
    costo_movimentazione_per_metro: Optional[float] = None
    costo_taccaggio_per_metro: Optional[float] = None
    costo_lavaggio_inizio_stagione: Optional[float] = None
    costo_lavaggio_fine_stagione: Optional[float] = None
    maggiorazione_scafo_sporco_per_metro: Optional[float] = None


class Cliente(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    cognome: str
    tipo_barca: str
    lunghezza: float  # metri
    tipo_sosta: str  # "dentro" | "fuori" | "fuori_sede" | "temporanea"
    giorni_sosta_temporanea: int = 0
    anno: int = Field(default_factory=lambda: datetime.now().year)
    posto_barca: Optional[int] = None  # 1-200
    telefono: Optional[str] = ""
    email: Optional[str] = ""
    codice_fiscale: Optional[str] = ""
    indirizzo: Optional[str] = ""
    cellulare: Optional[str] = ""
    # Pagamento
    pagato: bool = False
    data_pagamento: Optional[str] = None  # ISO date string
    # Motore
    potenza_motore: float = 0.0  # HP (cavalli)
    litri_olio_motore: float = 3.0
    litri_olio_piede: float = 1.0
    numero_candele: int = 4
    numero_termostati: int = 1
    # Secondo motore (opzionale)
    secondo_motore: bool = False
    potenza_motore_2: float = 0.0
    litri_olio_motore_2: float = 3.0
    litri_olio_piede_2: float = 1.0
    numero_candele_2: int = 4
    numero_termostati_2: int = 1
    girante_2_attivo: bool = True
    # Interruttori applicabilità
    antivegetativa_attiva: bool = True
    scafo_sporco_attivo: bool = False
    copertura_attiva: bool = False
    girante_attivo: bool = True
    lavaggio_inizio_attivo: bool = True
    lavaggio_fine_attivo: bool = True
    # Alaggio/Varo: spunta indipendente (attiva anche per soste dentro/temporanea/fuori_sede)
    alaggio_varo_attivo: bool = False
    # Destinazione alaggio/varo: "marina_di_campo" (tariffa fissa) o "altra" (costo manuale)
    destinazione_alaggio_varo: str = "marina_di_campo"
    destinazione_altra_nome: Optional[str] = ""
    # Costi (auto o manuali)
    costo_sosta: float = 0.0
    costo_copertura: float = 0.0
    costo_alaggio: float = 0.0
    costo_varo: float = 0.0
    costo_antivegetativa: float = 0.0
    costo_manutenzione_motore: float = 0.0
    costo_lavaggio_inizio: float = 0.0
    costo_lavaggio_fine: float = 0.0
    costo_scafo_sporco: float = 0.0
    costo_movimentazione: float = 0.0
    costo_taccaggio: float = 0.0
    # Breakdown ricambi (informativo)
    costo_ricambi_totale: float = 0.0
    costo_manodopera_motore: float = 0.0
    costo_ricambi_motore_2_totale: float = 0.0
    costo_manodopera_motore_2: float = 0.0
    # Lavorazioni extra (max 20): [{descrizione: str, prezzo: float}]
    lavorazioni_extra: List[dict] = Field(default_factory=list)
    # Override flags: se true, valore non ricalcolato automaticamente
    override_costi: bool = False
    # Lavori
    note_lavori: str = ""
    scadenza_antivegetativa: Optional[str] = None  # ISO date string
    scadenza_manutenzione: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ClienteCreate(BaseModel):
    nome: str
    cognome: str
    tipo_barca: str
    lunghezza: float
    tipo_sosta: str
    giorni_sosta_temporanea: Optional[int] = None
    anno: Optional[int] = None
    posto_barca: Optional[int] = None
    telefono: Optional[str] = ""
    email: Optional[str] = ""
    codice_fiscale: Optional[str] = ""
    indirizzo: Optional[str] = ""
    cellulare: Optional[str] = ""
    pagato: Optional[bool] = None
    data_pagamento: Optional[str] = None
    potenza_motore: Optional[float] = 0.0
    litri_olio_motore: Optional[float] = 3.0
    litri_olio_piede: Optional[float] = None
    numero_candele: Optional[int] = 4
    numero_termostati: Optional[int] = 1
    secondo_motore: Optional[bool] = False
    potenza_motore_2: Optional[float] = 0.0
    litri_olio_motore_2: Optional[float] = 3.0
    litri_olio_piede_2: Optional[float] = None
    numero_candele_2: Optional[int] = 4
    numero_termostati_2: Optional[int] = 1
    girante_2_attivo: Optional[bool] = None
    antivegetativa_attiva: Optional[bool] = True
    scafo_sporco_attivo: Optional[bool] = None
    copertura_attiva: Optional[bool] = None
    girante_attivo: Optional[bool] = True
    lavaggio_inizio_attivo: Optional[bool] = True
    lavaggio_fine_attivo: Optional[bool] = True
    costo_sosta: Optional[float] = None
    costo_copertura: Optional[float] = None
    costo_alaggio: Optional[float] = None
    costo_varo: Optional[float] = None
    alaggio_varo_attivo: Optional[bool] = None
    destinazione_alaggio_varo: Optional[str] = None
    destinazione_altra_nome: Optional[str] = None
    costo_antivegetativa: Optional[float] = None
    costo_manutenzione_motore: Optional[float] = None
    costo_lavaggio_inizio: Optional[float] = None
    costo_lavaggio_fine: Optional[float] = None
    costo_scafo_sporco: Optional[float] = None
    costo_movimentazione: Optional[float] = None
    costo_taccaggio: Optional[float] = None
    lavorazioni_extra: Optional[List[dict]] = None
    override_costi: bool = False
    note_lavori: str = ""
    scadenza_antivegetativa: Optional[str] = None
    scadenza_manutenzione: Optional[str] = None


# ---------- HELPERS ----------

def serialize(obj: BaseModel) -> dict:
    d = obj.model_dump()
    for k, v in d.items():
        if isinstance(v, datetime):
            d[k] = v.isoformat()
    return d


def deserialize_cliente(doc: dict) -> dict:
    if doc is None:
        return None
    doc.pop('_id', None)
    for k in ('created_at', 'updated_at'):
        if isinstance(doc.get(k), str):
            try:
                doc[k] = datetime.fromisoformat(doc[k])
            except Exception:
                pass
    return doc


def _sanitize_lavorazioni_extra(lst) -> List[dict]:
    """Valida e normalizza la lista lavorazioni extra: max 20 voci, ciascuna {descrizione, prezzo}."""
    if not lst:
        return []
    if not isinstance(lst, list):
        raise HTTPException(400, "lavorazioni_extra deve essere una lista")
    if len(lst) > 20:
        raise HTTPException(400, "Massimo 20 lavorazioni extra per cliente")
    out = []
    for item in lst:
        if not isinstance(item, dict):
            continue
        desc = str(item.get("descrizione") or "").strip()
        try:
            prezzo = round(float(item.get("prezzo") or 0), 2)
        except (TypeError, ValueError):
            prezzo = 0.0
        if not desc and prezzo == 0:
            continue
        out.append({"descrizione": desc, "prezzo": prezzo})
    return out


def _totale_extra(doc: dict) -> float:
    lst = doc.get("lavorazioni_extra") or []
    if not isinstance(lst, list):
        return 0.0
    return round(sum(float((it or {}).get("prezzo") or 0) for it in lst), 2)


async def get_tariffe_doc() -> Tariffe:
    doc = await db.tariffe.find_one({"id": "default"}, {"_id": 0})
    if not doc:
        t = Tariffe()
        await db.tariffe.insert_one(serialize(t))
        return t
    return Tariffe(**doc)


def calcola_alaggio(lunghezza: float, t: Tariffe) -> float:
    """Alaggio: forfait ≤5m e forfait >5m (due tariffe fisse, non moltiplicate per metro)."""
    if lunghezza <= 5:
        return round(t.alaggio_fino_5m, 2)
    return round(t.alaggio_oltre_5m_per_metro, 2)


def calcola_varo(lunghezza: float, t: Tariffe) -> float:
    """Varo: forfait ≤5m e forfait >5m (due tariffe fisse, non moltiplicate per metro)."""
    if lunghezza <= 5:
        return round(t.varo_fino_5m, 2)
    return round(t.varo_oltre_5m_per_metro, 2)


def calcola_motore_labor(potenza_hp: float, t: Tariffe) -> float:
    """Manodopera manutenzione motore in base a potenza HP.
    Scaglioni: 2-15 HP · 16-40 HP · 41-150 HP · oltre 150 HP.
    """
    if potenza_hp <= 0:
        return 0.0
    if potenza_hp <= 15:
        return round(t.motore_labor_2_15hp, 2)
    if potenza_hp <= 40:
        return round(t.motore_labor_fino_40hp, 2)
    if potenza_hp <= 150:
        return round(t.motore_labor_40_150hp, 2)
    return round(t.motore_labor_oltre_150hp, 2)


def calcola_ricambi(numero_candele: int, numero_termostati: int, t: Tariffe,
                    girante_attivo: bool = True, litri_olio_motore: float = 3.0,
                    litri_olio_piede: float = 1.0) -> dict:
    """Costo ricambi motore: girante, olio motore (× litri), filtro olio, candele, termostati, olio piede (× litri), anodi, ingrassaggio."""
    nc = int(numero_candele or 0)
    nt = int(numero_termostati or 0)
    litri = float(litri_olio_motore or 0)
    litri_piede = float(litri_olio_piede or 0)
    return {
        "girante": round(t.costo_girante, 2) if girante_attivo else 0.0,
        "olio_motore": round(litri * t.costo_olio_motore, 2),
        "filtro_olio": round(t.costo_filtro_olio, 2),
        "candele": round(nc * t.costo_candela, 2),
        "termostati": round(nt * t.costo_termostato, 2),
        "olio_piede": round(litri_piede * t.costo_olio_piede, 2),
        "anodi_interni": round(t.costo_anodi_interni, 2),
        "anodi_esterni": round(t.costo_anodi_esterni, 2),
        "ingrassaggio": round(t.costo_ingrassaggio, 2),
    }


def calcola_costi(lunghezza: float, tipo_sosta: str, t: Tariffe,
                  potenza_motore: float = 0.0, numero_candele: int = 4,
                  numero_termostati: int = 1,
                  antivegetativa_attiva: bool = True,
                  girante_attivo: bool = True,
                  litri_olio_motore: float = 3.0,
                  lavaggio_inizio_attivo: bool = True,
                  lavaggio_fine_attivo: bool = True,
                  secondo_motore: bool = False,
                  potenza_motore_2: float = 0.0,
                  litri_olio_motore_2: float = 3.0,
                  numero_candele_2: int = 4,
                  numero_termostati_2: int = 1,
                  girante_2_attivo: bool = True,
                  scafo_sporco_attivo: bool = False,
                  copertura_attiva: bool = False,
                  litri_olio_piede: float = 1.0,
                  litri_olio_piede_2: float = 1.0,
                  giorni_sosta_temporanea: int = 0,
                  destinazione_alaggio_varo: str = "marina_di_campo",
                  alaggio_varo_attivo: bool = False) -> dict:
    """Calcola costi automatici in base a lunghezza, tipo sosta e (uno o due) motori."""
    manodopera = calcola_motore_labor(potenza_motore, t)
    ricambi = calcola_ricambi(numero_candele, numero_termostati, t, girante_attivo, litri_olio_motore, litri_olio_piede)
    ricambi_tot = round(sum(ricambi.values()), 2)

    # Secondo motore (se presente)
    manodopera_2 = 0.0
    ricambi_2_tot = 0.0
    ricambi_2 = {}
    if secondo_motore:
        manodopera_2 = calcola_motore_labor(potenza_motore_2, t)
        ricambi_2 = calcola_ricambi(numero_candele_2, numero_termostati_2, t, girante_2_attivo, litri_olio_motore_2, litri_olio_piede_2)
        ricambi_2_tot = round(sum(ricambi_2.values()), 2)

    motore_tot = round(manodopera + ricambi_tot + manodopera_2 + ricambi_2_tot, 2)

    antiveg = round(lunghezza * t.antivegetativa_per_metro, 2) if antivegetativa_attiva else 0.0
    scafo_sporco = round(lunghezza * t.maggiorazione_scafo_sporco_per_metro, 2) if scafo_sporco_attivo else 0.0
    lav_inizio = round(lunghezza * t.costo_lavaggio_inizio_stagione, 2) if lavaggio_inizio_attivo else 0.0
    lav_fine = round(lunghezza * t.costo_lavaggio_fine_stagione, 2) if lavaggio_fine_attivo else 0.0

    base = {
        "costo_antivegetativa": antiveg,
        "costo_manutenzione_motore": motore_tot,
        "costo_manodopera_motore": manodopera,
        "costo_ricambi_totale": ricambi_tot,
        "costo_manodopera_motore_2": manodopera_2,
        "costo_ricambi_motore_2_totale": ricambi_2_tot,
        "costo_lavaggio_inizio": lav_inizio,
        "costo_lavaggio_fine": lav_fine,
        "costo_scafo_sporco": scafo_sporco,
        "ricambi_dettaglio": ricambi,
        "ricambi_2_dettaglio": ricambi_2,
    }

    # Movimentazione e taccaggio applicati solo per sosta fuori sede
    movimentazione = round(lunghezza * t.costo_movimentazione_per_metro, 2) if tipo_sosta == "fuori_sede" else 0.0
    taccaggio = round(lunghezza * t.costo_taccaggio_per_metro, 2) if tipo_sosta == "fuori_sede" else 0.0
    base["costo_movimentazione"] = movimentazione
    base["costo_taccaggio"] = taccaggio

    # Copertura: ora è spunta indipendente, non più automatica per sosta fuori
    copertura = round(lunghezza * t.copertura_per_metro, 2) if copertura_attiva else 0.0

    # Alaggio/Varo: spunta indipendente. Se attiva, calcola sempre (tranne se destinazione="altra" = manuale)
    if alaggio_varo_attivo:
        if destinazione_alaggio_varo == "marina_di_campo":
            alaggio_val = calcola_alaggio(lunghezza, t)
            varo_val = calcola_varo(lunghezza, t)
        else:
            alaggio_val = 0.0
            varo_val = 0.0
    else:
        alaggio_val = 0.0
        varo_val = 0.0

    if tipo_sosta == "fuori":
        base.update({
            "costo_sosta": round(lunghezza * t.sosta_fuori_per_metro, 2),
            "costo_copertura": copertura,
            "costo_alaggio": alaggio_val,
            "costo_varo": varo_val,
        })
    elif tipo_sosta == "fuori_sede":
        # Sosta fuori sede: la barca è custodita dal cliente. Copertura/alaggio/varo opzionali.
        base.update({
            "costo_sosta": 0.0,
            "costo_copertura": copertura,
            "costo_alaggio": alaggio_val,
            "costo_varo": varo_val,
        })
    elif tipo_sosta == "temporanea":
        # Tariffa a giornata × numero di giorni
        giorni = int(giorni_sosta_temporanea or 0)
        base.update({
            "costo_sosta": round(giorni * t.sosta_temporanea_giornaliera, 2),
            "costo_copertura": copertura,
            "costo_alaggio": alaggio_val,
            "costo_varo": varo_val,
        })
    else:
        base.update({
            "costo_sosta": round(lunghezza * t.sosta_dentro_per_metro, 2),
            "costo_copertura": copertura,
            "costo_alaggio": alaggio_val,
            "costo_varo": varo_val,
        })
    return base


# ---------- ROUTES ----------

@api_router.get("/")
async def root():
    return {"message": "Cantiere Nautico API - OK"}


# --- Tariffe ---
@api_router.get("/tariffe", response_model=Tariffe)
async def get_tariffe():
    return await get_tariffe_doc()


@api_router.put("/tariffe", response_model=Tariffe)
async def update_tariffe(payload: TariffeUpdate):
    current = await get_tariffe_doc()
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    new_data = current.model_dump()
    new_data.update(updates)
    new_data["updated_at"] = datetime.now(timezone.utc)
    t = Tariffe(**new_data)
    await db.tariffe.update_one({"id": "default"}, {"$set": serialize(t)}, upsert=True)
    return t


@api_router.post("/tariffe/ricalcola")
async def ricalcola_costi_anno(anno: int):
    """Ricalcola i costi di tutti i clienti dell'anno indicato usando le tariffe correnti.
    Rispetta: override_costi (non ricalcola i costi manuali) e destinazione="altra" (preserva alaggio/varo manuali)."""
    t = await get_tariffe_doc()
    clienti_docs = await db.clienti.find({"anno": anno}, {"_id": 0}).to_list(10000)
    aggiornati = 0
    for c in clienti_docs:
        try:
            auto_costi = calcola_costi(
                float(c.get("lunghezza") or 0),
                str(c.get("tipo_sosta") or "dentro"),
                t,
                float(c.get("potenza_motore") or 0),
                int(c.get("numero_candele") or 4),
                int(c.get("numero_termostati") or 1),
                bool(c.get("antivegetativa_attiva", True)),
                bool(c.get("girante_attivo", True)),
                float(c.get("litri_olio_motore") or 3.0),
                bool(c.get("lavaggio_inizio_attivo", True)),
                bool(c.get("lavaggio_fine_attivo", True)),
                bool(c.get("secondo_motore", False)),
                float(c.get("potenza_motore_2") or 0),
                float(c.get("litri_olio_motore_2") or 3.0),
                int(c.get("numero_candele_2") or 4),
                int(c.get("numero_termostati_2") or 1),
                bool(c.get("girante_2_attivo", True)),
                bool(c.get("scafo_sporco_attivo", False)),
                bool(c.get("copertura_attiva", False)),
                float(c.get("litri_olio_piede") or 1.0),
                float(c.get("litri_olio_piede_2") or 1.0),
                int(c.get("giorni_sosta_temporanea") or 0),
                str(c.get("destinazione_alaggio_varo") or "marina_di_campo"),
                bool(c.get("alaggio_varo_attivo", False)),
            )
            auto_costi.pop("ricambi_dettaglio", None)
            auto_costi.pop("ricambi_2_dettaglio", None)

            override = bool(c.get("override_costi", False))
            manual_av = bool(c.get("alaggio_varo_attivo", False)) and str(c.get("destinazione_alaggio_varo")) == "altra"

            updates = {}
            for k, v in auto_costi.items():
                if override:
                    continue  # non toccare i costi manuali
                if manual_av and k in ("costo_alaggio", "costo_varo"):
                    continue  # preserva valori manuali per destinazione altra
                updates[k] = v
            if updates:
                updates["updated_at"] = datetime.now(timezone.utc).isoformat()
                await db.clienti.update_one({"id": c["id"]}, {"$set": updates})
                aggiornati += 1
        except Exception as e:
            logger.warning(f"Ricalcolo cliente {c.get('id')} fallito: {e}")
    return {"ok": True, "anno": anno, "aggiornati": aggiornati, "totali": len(clienti_docs)}


@api_router.get("/tariffe/listino.pdf")
async def listino_prezzi_pdf():
    """Genera il listino prezzi ufficiale su carta intestata del cantiere.
    Include tutte le tariffe raggruppate per categoria + data emissione ben visibile.
    """
    t = await get_tariffe_doc()
    cantiere = await db.cantiere.find_one({"id": "default"}, {"_id": 0}) or {}

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm,
        title=f"Listino prezzi — {cantiere.get('nome') or 'Cantiere Nautico'}"
    )
    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#0F1B3D")
    TEAK = colors.HexColor("#B0562E")
    SAND = colors.HexColor("#F3EFE7")
    MUTED = colors.HexColor("#5B6478")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=18, textColor=NAVY, spaceAfter=1, leading=20)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=9, textColor=TEAK, spaceBefore=3, spaceAfter=1, leading=11)
    body = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica", fontSize=8, textColor=NAVY, leading=10)
    tiny = ParagraphStyle("tiny", parent=styles["Normal"], fontName="Helvetica", fontSize=7, textColor=MUTED, leading=9)
    date_style = ParagraphStyle("date", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, textColor=TEAK, leading=12)

    elems = []
    # Header carta intestata
    nome_cantiere = (cantiere.get("nome") or "CANTIERE NAUTICO").upper()
    contatti = " · ".join([x for x in [
        cantiere.get("indirizzo"),
        cantiere.get("telefono"),
        cantiere.get("email"),
        cantiere.get("piva") and f"P.IVA {cantiere['piva']}",
    ] if x])

    # Header con eventuale logo affiancato al titolo
    logo_b64 = cantiere.get("logo_base64")
    left_col = [Paragraph(f"<b>{nome_cantiere}</b>", h1)]
    if cantiere.get("slogan"):
        left_col.append(Paragraph(f"<font color='#5B6478' size=8><i>{cantiere['slogan']}</i></font>", body))
    if contatti:
        left_col.append(Spacer(1, 1*mm))
        left_col.append(Paragraph(f"<font color='#5B6478' size=8>{contatti}</font>", body))

    if logo_b64:
        try:
            import base64 as _b64
            raw = _b64.b64decode(logo_b64.split(",")[-1] if "," in logo_b64 else logo_b64)
            img = Image(io.BytesIO(raw), width=22*mm, height=22*mm, kind="proportional")
            header_tbl = Table([[left_col, img]], colWidths=[150*mm, 30*mm])
        except Exception:
            header_tbl = Table([[left_col, ""]], colWidths=[150*mm, 30*mm])
    else:
        header_tbl = Table([[left_col, ""]], colWidths=[150*mm, 30*mm])
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (1,0), (1,0), "RIGHT"),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    elems.append(header_tbl)
    elems.append(Spacer(1, 2*mm))

    # Data + titolo listino
    sep = Table([[""]], colWidths=[180*mm], rowHeights=[1.5])
    sep.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), TEAK)]))
    elems.append(sep)
    elems.append(Spacer(1, 2*mm))

    title_tbl = Table([[
        Paragraph("<b>LISTINO PREZZI</b> · <font color='#5B6478' size=8>Tariffario in vigore</font>", h1),
        Paragraph(f"Data emissione · <font size=12 color='#B0562E'>{date.today().strftime('%d/%m/%Y')}</font>", date_style),
    ]], colWidths=[110*mm, 70*mm])
    title_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (1,0), (1,0), "RIGHT"),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    elems.append(title_tbl)
    elems.append(Spacer(1, 2*mm))

    # Gruppi tariffe
    groups = [
        ("SOSTA", [
            ("Sosta al coperto", "sosta_dentro_per_metro", "€ / metro / anno"),
            ("Sosta su piazzale (fuori)", "sosta_fuori_per_metro", "€ / metro / anno"),
            ("Sosta temporanea", "sosta_temporanea_giornaliera", "€ / giorno"),
            ("Movimentazione (fuori sede)", "costo_movimentazione_per_metro", "€ / metro"),
            ("Taccaggio (fuori sede)", "costo_taccaggio_per_metro", "€ / metro"),
        ]),
        ("ALAGGIO & VARO", [
            ("Alaggio · fino a 5 m", "alaggio_fino_5m", "forfait"),
            ("Alaggio · oltre 5 m", "alaggio_oltre_5m_per_metro", "forfait"),
            ("Varo · fino a 5 m", "varo_fino_5m", "forfait"),
            ("Varo · oltre 5 m", "varo_oltre_5m_per_metro", "forfait"),
        ]),
        ("COPERTURA & TRATTAMENTI SCAFO", [
            ("Copertura", "copertura_per_metro", "€ / metro"),
            ("Antivegetativa", "antivegetativa_per_metro", "€ / metro"),
            ("Maggiorazione scafo sporco", "maggiorazione_scafo_sporco_per_metro", "€ / metro"),
            ("Lavaggio inizio stagione", "costo_lavaggio_inizio_stagione", "€ / metro"),
            ("Lavaggio fine stagione", "costo_lavaggio_fine_stagione", "€ / metro"),
        ]),
        ("MANODOPERA MOTORE", [
            ("Manodopera · 2-15 HP", "motore_labor_2_15hp", "forfait"),
            ("Manodopera · 16-40 HP", "motore_labor_fino_40hp", "forfait"),
            ("Manodopera · 41-150 HP", "motore_labor_40_150hp", "forfait"),
            ("Manodopera · oltre 150 HP", "motore_labor_oltre_150hp", "forfait"),
        ]),
        ("RICAMBI & MATERIALI", [
            ("Girante", "costo_girante", "cad."),
            ("Olio motore", "costo_olio_motore", "€ / litro"),
            ("Filtro olio", "costo_filtro_olio", "cad."),
            ("Candela", "costo_candela", "cad."),
            ("Termostato", "costo_termostato", "cad."),
            ("Olio piede", "costo_olio_piede", "€ / litro"),
            ("Kit anodi interni", "costo_anodi_interni", "forfait"),
            ("Kit anodi esterni", "costo_anodi_esterni", "forfait"),
            ("Ingrassaggio completo", "costo_ingrassaggio", "forfait"),
        ]),
    ]

    tariffe_dict = t.model_dump()
    for titolo, voci in groups:
        elems.append(Paragraph(titolo, h2))
        rows = [["VOCE", "UNITÀ", "IMPORTO"]]
        for label, key, unit in voci:
            val = tariffe_dict.get(key, 0) or 0
            rows.append([label, unit, _euro(float(val))])
        tbl = Table(rows, colWidths=[100*mm, 40*mm, 40*mm])
        n = len(rows) - 1
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), NAVY),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 7),
            ("ALIGN", (2,0), (2,-1), "RIGHT"),
            ("ALIGN", (1,0), (1,-1), "CENTER"),
            ("FONTNAME", (0,1), (-1,n), "Helvetica"),
            ("FONTSIZE", (0,1), (-1,n), 8),
            ("TEXTCOLOR", (0,1), (-1,n), NAVY),
            ("ROWBACKGROUNDS", (0,1), (-1,n), [colors.white, SAND]),
            ("LINEBELOW", (0,1), (-1,n), 0.3, colors.HexColor("#D9D9D9")),
            ("TOPPADDING", (0,0), (-1,-1), 1),
            ("BOTTOMPADDING", (0,0), (-1,-1), 1),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ]))
        elems.append(tbl)

    elems.append(Spacer(1, 2*mm))
    elems.append(Paragraph(
        f"Prezzi al netto di IVA salvo diversa indicazione · Listino aggiornato al {date.today().strftime('%d/%m/%Y')} · "
        f"Per preventivi personalizzati contattare {cantiere.get('telefono') or cantiere.get('email') or 'il cantiere'}.",
        tiny
    ))

    pdf.build(elems)
    buf.seek(0)
    filename = f"listino_prezzi_{date.today().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# --- Preview costi ---
@api_router.get("/calcola-costi")
async def preview_costi(lunghezza: float, tipo_sosta: str,
                        potenza_motore: float = 0.0,
                        numero_candele: int = 4,
                        numero_termostati: int = 1,
                        antivegetativa_attiva: bool = True,
                        girante_attivo: bool = True,
                        litri_olio_motore: float = 3.0,
                        lavaggio_inizio_attivo: bool = True,
                        lavaggio_fine_attivo: bool = True,
                        secondo_motore: bool = False,
                        potenza_motore_2: float = 0.0,
                        litri_olio_motore_2: float = 3.0,
                        numero_candele_2: int = 4,
                        numero_termostati_2: int = 1,
                        girante_2_attivo: bool = True,
                        scafo_sporco_attivo: bool = False,
                        copertura_attiva: bool = False,
                        litri_olio_piede: float = 1.0,
                        litri_olio_piede_2: float = 1.0,
                        giorni_sosta_temporanea: int = 0,
                        destinazione_alaggio_varo: str = "marina_di_campo",
                        alaggio_varo_attivo: bool = False):
    if tipo_sosta not in ("dentro", "fuori", "fuori_sede", "temporanea"):
        raise HTTPException(400, "tipo_sosta deve essere 'dentro', 'fuori', 'fuori_sede' o 'temporanea'")
    t = await get_tariffe_doc()
    return calcola_costi(lunghezza, tipo_sosta, t, potenza_motore,
                         numero_candele, numero_termostati,
                         antivegetativa_attiva, girante_attivo, litri_olio_motore,
                         lavaggio_inizio_attivo, lavaggio_fine_attivo,
                         secondo_motore, potenza_motore_2, litri_olio_motore_2,
                         numero_candele_2, numero_termostati_2, girante_2_attivo,
                         scafo_sporco_attivo, copertura_attiva,
                         litri_olio_piede, litri_olio_piede_2,
                         giorni_sosta_temporanea, destinazione_alaggio_varo,
                         alaggio_varo_attivo)


# --- Clienti ---
@api_router.get("/clienti", response_model=List[Cliente])
async def list_clienti(anno: Optional[int] = None):
    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(1000)
    # Ordina alfabeticamente per cognome (case-insensitive), poi nome
    docs.sort(key=lambda d: ((d.get("cognome") or "").strip().lower(), (d.get("nome") or "").strip().lower()))
    return [Cliente(**deserialize_cliente(d)) for d in docs]


@api_router.get("/clienti/{cliente_id}", response_model=Cliente)
async def get_cliente(cliente_id: str):
    doc = await db.clienti.find_one({"id": cliente_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Cliente non trovato")
    return Cliente(**deserialize_cliente(doc))


@api_router.post("/clienti", response_model=Cliente)
async def create_cliente(payload: ClienteCreate):
    if payload.tipo_sosta not in ("dentro", "fuori", "fuori_sede"):
        raise HTTPException(400, "tipo_sosta deve essere 'dentro', 'fuori' o 'fuori_sede'")
    if payload.posto_barca is not None:
        if payload.posto_barca < 1 or payload.posto_barca > TOTAL_POSTI:
            raise HTTPException(400, f"Posto barca deve essere tra 1 e {TOTAL_POSTI}")
        anno_check = payload.anno or datetime.now().year
        existing = await db.clienti.find_one({"posto_barca": payload.posto_barca, "anno": anno_check})
        if existing:
            raise HTTPException(400, f"Posto barca {payload.posto_barca} già occupato per l'anno {anno_check}")

    t = await get_tariffe_doc()
    auto_costi = calcola_costi(
        payload.lunghezza, payload.tipo_sosta, t,
        payload.potenza_motore or 0,
        payload.numero_candele or 4,
        payload.numero_termostati or 1,
        bool(payload.antivegetativa_attiva if payload.antivegetativa_attiva is not None else True),
        bool(payload.girante_attivo if payload.girante_attivo is not None else True),
        float(payload.litri_olio_motore if payload.litri_olio_motore is not None else 3.0),
        bool(payload.lavaggio_inizio_attivo if payload.lavaggio_inizio_attivo is not None else True),
        bool(payload.lavaggio_fine_attivo if payload.lavaggio_fine_attivo is not None else True),
        bool(payload.secondo_motore if payload.secondo_motore is not None else False),
        float(payload.potenza_motore_2 or 0),
        float(payload.litri_olio_motore_2 if payload.litri_olio_motore_2 is not None else 3.0),
        int(payload.numero_candele_2 or 4),
        int(payload.numero_termostati_2 or 1),
        bool(payload.girante_2_attivo if payload.girante_2_attivo is not None else True),
        bool(payload.scafo_sporco_attivo if payload.scafo_sporco_attivo is not None else False),
        bool(payload.copertura_attiva if payload.copertura_attiva is not None else False),
        float(payload.litri_olio_piede if payload.litri_olio_piede is not None else 1.0),
        float(payload.litri_olio_piede_2 if payload.litri_olio_piede_2 is not None else 1.0),
        int(payload.giorni_sosta_temporanea or 0),
        (payload.destinazione_alaggio_varo or "marina_di_campo"),
        bool(payload.alaggio_varo_attivo if payload.alaggio_varo_attivo is not None else False),
    )
    # Rimuovi dettaglio non-modello prima di applicarlo
    ricambi_dettaglio = auto_costi.pop("ricambi_dettaglio", None)
    ricambi_2_dettaglio = auto_costi.pop("ricambi_2_dettaglio", None)

    data = payload.model_dump()
    # Sanitize lavorazioni_extra (max 20, normalize)
    if data.get("lavorazioni_extra") is not None:
        data["lavorazioni_extra"] = _sanitize_lavorazioni_extra(data["lavorazioni_extra"])
    # Destinazione "altra" + alaggio_varo_attivo: costi manuali (bypass override)
    manual_alaggio_varo = (payload.alaggio_varo_attivo and payload.destinazione_alaggio_varo == "altra")
    # Se override non attivo → usa costi calcolati; altrimenti usa quelli inseriti (fallback 0 se None)
    for k in auto_costi:
        val = data.get(k)
        if manual_alaggio_varo and k in ("costo_alaggio", "costo_varo"):
            data[k] = float(val) if val is not None else 0.0
        elif not payload.override_costi or val is None:
            data[k] = auto_costi[k]
        else:
            data[k] = val

    cliente = Cliente(**{k: v for k, v in data.items() if v is not None or k in ("posto_barca", "scadenza_antivegetativa", "scadenza_manutenzione")})
    await db.clienti.insert_one(serialize(cliente))
    return cliente


@api_router.put("/clienti/{cliente_id}", response_model=Cliente)
async def update_cliente(cliente_id: str, payload: ClienteCreate):
    existing = await db.clienti.find_one({"id": cliente_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Cliente non trovato")

    if payload.tipo_sosta not in ("dentro", "fuori", "fuori_sede"):
        raise HTTPException(400, "tipo_sosta deve essere 'dentro', 'fuori' o 'fuori_sede'")

    if payload.posto_barca is not None:
        if payload.posto_barca < 1 or payload.posto_barca > TOTAL_POSTI:
            raise HTTPException(400, f"Posto barca deve essere tra 1 e {TOTAL_POSTI}")
        anno_check = payload.anno or existing.get("anno") or datetime.now().year
        conflict = await db.clienti.find_one({"posto_barca": payload.posto_barca, "anno": anno_check, "id": {"$ne": cliente_id}})
        if conflict:
            raise HTTPException(400, f"Posto barca {payload.posto_barca} già occupato per l'anno {anno_check}")

    t = await get_tariffe_doc()
    auto_costi = calcola_costi(
        payload.lunghezza, payload.tipo_sosta, t,
        payload.potenza_motore or 0,
        payload.numero_candele or 4,
        payload.numero_termostati or 1,
        bool(payload.antivegetativa_attiva if payload.antivegetativa_attiva is not None else True),
        bool(payload.girante_attivo if payload.girante_attivo is not None else True),
        float(payload.litri_olio_motore if payload.litri_olio_motore is not None else 3.0),
        bool(payload.lavaggio_inizio_attivo if payload.lavaggio_inizio_attivo is not None else True),
        bool(payload.lavaggio_fine_attivo if payload.lavaggio_fine_attivo is not None else True),
        bool(payload.secondo_motore if payload.secondo_motore is not None else False),
        float(payload.potenza_motore_2 or 0),
        float(payload.litri_olio_motore_2 if payload.litri_olio_motore_2 is not None else 3.0),
        int(payload.numero_candele_2 or 4),
        int(payload.numero_termostati_2 or 1),
        bool(payload.girante_2_attivo if payload.girante_2_attivo is not None else True),
        bool(payload.scafo_sporco_attivo if payload.scafo_sporco_attivo is not None else False),
        bool(payload.copertura_attiva if payload.copertura_attiva is not None else False),
        float(payload.litri_olio_piede if payload.litri_olio_piede is not None else 1.0),
        float(payload.litri_olio_piede_2 if payload.litri_olio_piede_2 is not None else 1.0),
        int(payload.giorni_sosta_temporanea or 0),
        (payload.destinazione_alaggio_varo or existing.get("destinazione_alaggio_varo") or "marina_di_campo"),
        bool(payload.alaggio_varo_attivo if payload.alaggio_varo_attivo is not None else existing.get("alaggio_varo_attivo", False)),
    )
    auto_costi.pop("ricambi_dettaglio", None)
    auto_costi.pop("ricambi_2_dettaglio", None)

    data = payload.model_dump()
    # Sanitize lavorazioni_extra (max 20, normalize)
    if data.get("lavorazioni_extra") is not None:
        data["lavorazioni_extra"] = _sanitize_lavorazioni_extra(data["lavorazioni_extra"])
    # Destinazione "altra" + alaggio_varo_attivo: alaggio/varo manuali
    dest_effettiva = payload.destinazione_alaggio_varo or existing.get("destinazione_alaggio_varo") or "marina_di_campo"
    av_attivo = payload.alaggio_varo_attivo if payload.alaggio_varo_attivo is not None else existing.get("alaggio_varo_attivo", False)
    manual_alaggio_varo = (av_attivo and dest_effettiva == "altra")
    for k in auto_costi:
        val = data.get(k)
        if manual_alaggio_varo and k in ("costo_alaggio", "costo_varo"):
            data[k] = float(val) if val is not None else 0.0
        elif not payload.override_costi or val is None:
            data[k] = auto_costi[k]
        else:
            data[k] = val

    merged = deserialize_cliente(existing)
    merged.update({k: v for k, v in data.items() if v is not None or k in ("posto_barca", "scadenza_antivegetativa", "scadenza_manutenzione")})
    merged["id"] = cliente_id
    merged["updated_at"] = datetime.now(timezone.utc)
    cliente = Cliente(**merged)
    await db.clienti.update_one({"id": cliente_id}, {"$set": serialize(cliente)})
    return cliente


@api_router.delete("/clienti/{cliente_id}")
async def delete_cliente(cliente_id: str):
    res = await db.clienti.delete_one({"id": cliente_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Cliente non trovato")
    return {"ok": True}


# --- Stats ---
@api_router.get("/stats")
async def stats(anno: Optional[int] = None):
    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(1000)
    dentro = sum(1 for d in docs if d.get("tipo_sosta") == "dentro")
    fuori = sum(1 for d in docs if d.get("tipo_sosta") == "fuori")
    fuori_sede = sum(1 for d in docs if d.get("tipo_sosta") == "fuori_sede")
    occupati = sum(1 for d in docs if d.get("posto_barca"))
    liberi = TOTAL_POSTI - occupati

    entrate_totali = 0.0
    for d in docs:
        entrate_totali += sum([
            d.get("costo_sosta", 0) or 0,
            d.get("costo_movimentazione", 0) or 0,
            d.get("costo_taccaggio", 0) or 0,
            d.get("costo_copertura", 0) or 0,
            d.get("costo_alaggio", 0) or 0,
            d.get("costo_varo", 0) or 0,
            d.get("costo_antivegetativa", 0) or 0,
            d.get("costo_manutenzione_motore", 0) or 0,
            d.get("costo_lavaggio_inizio", 0) or 0,
            d.get("costo_lavaggio_fine", 0) or 0,
            d.get("costo_scafo_sporco", 0) or 0,
        ])
        entrate_totali += _totale_extra(d)

    # Scadenze prossime (entro 30 giorni)
    from datetime import timedelta
    oggi = date.today()
    limite = oggi + timedelta(days=30)
    scadenze = []
    for d in docs:
        for tipo, key in (("Antivegetativa", "scadenza_antivegetativa"), ("Manutenzione motore", "scadenza_manutenzione")):
            v = d.get(key)
            if v:
                try:
                    dt = datetime.fromisoformat(v).date() if "T" in v else date.fromisoformat(v)
                    if oggi <= dt <= limite:
                        scadenze.append({
                            "cliente_id": d["id"],
                            "nome": f"{d.get('nome','')} {d.get('cognome','')}",
                            "tipo": tipo,
                            "data": dt.isoformat(),
                            "giorni_rimanenti": (dt - oggi).days,
                        })
                except Exception:
                    pass
    scadenze.sort(key=lambda x: x["giorni_rimanenti"])

    return {
        "totale_clienti": len(docs),
        "posti_totali": TOTAL_POSTI,
        "posti_occupati": occupati,
        "posti_liberi": liberi,
        "sosta_dentro": dentro,
        "sosta_fuori": fuori,
        "sosta_fuori_sede": fuori_sede,
        "entrate_totali": round(entrate_totali, 2),
        "scadenze_prossime": scadenze[:10],
    }


# --- Posti barca ---
@api_router.get("/posti-barca")
async def posti_barca(anno: Optional[int] = None):
    q = {"posto_barca": {"$ne": None}}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(1000)
    occupati_map = {d["posto_barca"]: d for d in docs if d.get("posto_barca")}
    result = []
    for i in range(1, TOTAL_POSTI + 1):
        c = occupati_map.get(i)
        result.append({
            "numero": i,
            "occupato": c is not None,
            "cliente_id": c.get("id") if c else None,
            "cliente_nome": f"{c.get('nome','')} {c.get('cognome','')}" if c else None,
            "tipo_sosta": c.get("tipo_sosta") if c else None,
            "tipo_barca": c.get("tipo_barca") if c else None,
        })
    return result


@api_router.get("/posti-barca/next")
async def next_posto_libero(anno: Optional[int] = None, escludi_cliente_id: Optional[str] = None):
    """Ritorna il primo posto barca libero (1-200) per l'anno indicato.
    Se `escludi_cliente_id` è passato, il posto attualmente occupato da quel cliente è considerato libero
    (utile in fase di modifica per non "bloccare" il posto già assegnato al cliente stesso).
    """
    anno_target = anno if anno is not None else datetime.now().year
    q = {"posto_barca": {"$ne": None}, "anno": anno_target}
    if escludi_cliente_id:
        q["id"] = {"$ne": escludi_cliente_id}
    docs = await db.clienti.find(q, {"posto_barca": 1, "_id": 0}).to_list(2000)
    occupati = {int(d["posto_barca"]) for d in docs if d.get("posto_barca")}
    for i in range(1, TOTAL_POSTI + 1):
        if i not in occupati:
            return {"anno": anno_target, "posto": i, "posti_liberi": TOTAL_POSTI - len(occupati)}
    return {"anno": anno_target, "posto": None, "posti_liberi": 0}


# --- Export ---
@api_router.get("/export/clienti.csv")
async def export_csv():
    docs = await db.clienti.find({}, {"_id": 0}).to_list(1000)
    if not docs:
        docs = [{}]
    df = pd.DataFrame(docs)
    if not df.empty:
        cols_order = ["posto_barca", "nome", "cognome", "tipo_barca", "lunghezza", "tipo_sosta",
                      "telefono", "email", "costo_sosta", "costo_copertura", "costo_alaggio",
                      "costo_varo", "costo_antivegetativa", "costo_manutenzione_motore",
                      "scadenza_antivegetativa", "scadenza_manutenzione", "note_lavori"]
        cols = [c for c in cols_order if c in df.columns] + [c for c in df.columns if c not in cols_order]
        df = df[cols]

    buf = io.StringIO()
    df.to_csv(buf, index=False, sep=";")
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clienti_cantiere.csv"}
    )


@api_router.get("/export/clienti.xlsx")
async def export_xlsx(anno: Optional[int] = None):
    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(1000)
    docs.sort(key=lambda d: ((d.get("cognome") or "").strip().lower(), (d.get("nome") or "").strip().lower()))

    # Colonne human-readable per il commercialista
    COLS = [
        ("Anno", "anno"),
        ("Posto", "posto_barca"),
        ("Cognome", "cognome"),
        ("Nome", "nome"),
        ("Codice Fiscale", "codice_fiscale"),
        ("Indirizzo", "indirizzo"),
        ("Telefono", "telefono"),
        ("Cellulare", "cellulare"),
        ("Email", "email"),
        ("Tipo barca", "tipo_barca"),
        ("Lunghezza (m)", "lunghezza"),
        ("Tipo sosta", "tipo_sosta"),
        ("Gg. sosta temp.", "giorni_sosta_temporanea"),
        ("Destinazione alaggio/varo", "destinazione_alaggio_varo"),
        ("Nome altra destinazione", "destinazione_altra_nome"),
        ("Sosta €", "costo_sosta"),
        ("Movimentazione €", "costo_movimentazione"),
        ("Taccaggio €", "costo_taccaggio"),
        ("Copertura €", "costo_copertura"),
        ("Alaggio €", "costo_alaggio"),
        ("Varo €", "costo_varo"),
        ("Antivegetativa €", "costo_antivegetativa"),
        ("Magg. scafo sporco €", "costo_scafo_sporco"),
        ("Lavaggio inizio €", "costo_lavaggio_inizio"),
        ("Lavaggio fine €", "costo_lavaggio_fine"),
        ("Manutenzione motore €", "costo_manutenzione_motore"),
        ("Lavorazioni extra €", "__totale_extra__"),
        ("TOTALE €", "__totale__"),
        ("Pagato", "__pagato__"),
        ("Data pagamento", "data_pagamento"),
        ("Scad. antivegetativa", "scadenza_antivegetativa"),
        ("Scad. manutenzione", "scadenza_manutenzione"),
        ("Note lavori", "note_lavori"),
    ]

    COST_KEYS = ("costo_sosta","costo_movimentazione","costo_taccaggio","costo_copertura",
                 "costo_alaggio","costo_varo","costo_antivegetativa","costo_scafo_sporco",
                 "costo_lavaggio_inizio","costo_lavaggio_fine","costo_manutenzione_motore")

    def row_for(d: dict):
        tot_extra = round(sum(float((it or {}).get("prezzo") or 0) for it in (d.get("lavorazioni_extra") or [])), 2)
        totale = round(sum(float(d.get(k) or 0) for k in COST_KEYS) + tot_extra, 2)
        sosta_map = {"dentro": "Al coperto", "fuori": "Su piazzale", "fuori_sede": "Fuori sede", "temporanea": "Temporanea"}
        dest_map = {"marina_di_campo": "Marina di Campo", "altra": "Altra"}
        out = {}
        for label, key in COLS:
            if key == "__totale__":
                out[label] = totale
            elif key == "__totale_extra__":
                out[label] = tot_extra
            elif key == "__pagato__":
                out[label] = "Sì" if d.get("pagato") else "No"
            elif key == "tipo_sosta":
                out[label] = sosta_map.get(d.get("tipo_sosta"), d.get("tipo_sosta") or "")
            elif key == "destinazione_alaggio_varo":
                out[label] = dest_map.get(d.get("destinazione_alaggio_varo"), "")
            else:
                out[label] = d.get(key, "")
        return out

    rows = [row_for(d) for d in docs]
    df = pd.DataFrame(rows, columns=[c[0] for c in COLS]) if rows else pd.DataFrame(columns=[c[0] for c in COLS])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = f"Clienti {anno}" if anno else "Clienti"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        # Formattazione: header grassetto, auto-width, valute a 2 decimali, riga totali
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color="17324D", end_color="17324D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        totali_fill = PatternFill(start_color="B0562E", end_color="B0562E", fill_type="solid")
        totali_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(border_style="thin", color="D9D9D9")

        # Header row
        for col_idx, (label, _) in enumerate(COLS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        # Colonne valuta (indice 1-based)
        currency_labels = {label for label, _ in COLS if label.endswith("€")}
        for row_idx in range(2, len(rows) + 2):
            for col_idx, (label, _) in enumerate(COLS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if label in currency_labels:
                    cell.number_format = '#,##0.00 "€"'
                    cell.alignment = Alignment(horizontal="right")
                elif label == "Lunghezza (m)":
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right")
                if label == "TOTALE €":
                    cell.font = Font(bold=True)

        # Riga TOTALI in fondo
        if rows:
            tot_row = len(rows) + 2
            ws.cell(row=tot_row, column=1, value="TOTALI").font = totali_font
            ws.cell(row=tot_row, column=1).fill = totali_fill
            ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="right")
            # Merge cells da colonna 1 a 15 per etichetta TOTALI
            ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=15)
            for col_idx, (label, _) in enumerate(COLS, start=1):
                if label in currency_labels:
                    col_letter = get_column_letter(col_idx)
                    ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{tot_row-1})")
                    ws.cell(row=tot_row, column=col_idx).number_format = '#,##0.00 "€"'
                    ws.cell(row=tot_row, column=col_idx).font = totali_font
                    ws.cell(row=tot_row, column=col_idx).fill = totali_fill
                    ws.cell(row=tot_row, column=col_idx).alignment = Alignment(horizontal="right")

        # Auto-width colonne
        widths = {
            "Anno": 8, "Posto": 8, "Cognome": 16, "Nome": 14, "Codice Fiscale": 20,
            "Indirizzo": 28, "Telefono": 14, "Cellulare": 14, "Email": 24,
            "Tipo barca": 18, "Lunghezza (m)": 12, "Tipo sosta": 14,
            "Gg. sosta temp.": 10, "Destinazione alaggio/varo": 22, "Nome altra destinazione": 22,
            "Pagato": 8, "Data pagamento": 14, "Scad. antivegetativa": 16, "Scad. manutenzione": 16,
            "Note lavori": 40,
        }
        for col_idx, (label, _) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(label, 15 if label.endswith("€") else 14)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    buf.seek(0)
    fname = f"clienti_cantiere_{anno}.xlsx" if anno else "clienti_cantiere.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ---------- LAVORI (storico strutturato) ----------

class Lavoro(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cliente_id: str
    data: str  # ISO date string YYYY-MM-DD
    tipo: str  # es. Antivegetativa, Manutenzione motore, Riparazione, Altro
    descrizione: str = ""
    costo: float = 0.0
    materiali: str = ""
    stato: str = "completato"  # pianificato | in_corso | completato
    anno: int = Field(default_factory=lambda: datetime.now().year)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LavoroCreate(BaseModel):
    cliente_id: str
    data: str
    tipo: str
    descrizione: Optional[str] = ""
    costo: Optional[float] = 0.0
    materiali: Optional[str] = ""
    stato: Optional[str] = "completato"
    anno: Optional[int] = None


@api_router.get("/clienti/{cliente_id}/lavori", response_model=List[Lavoro])
async def list_lavori(cliente_id: str):
    docs = await db.lavori.find({"cliente_id": cliente_id}, {"_id": 0}).sort("data", -1).to_list(1000)
    for d in docs:
        if isinstance(d.get("created_at"), str):
            try:
                d["created_at"] = datetime.fromisoformat(d["created_at"])
            except Exception:
                pass
    return [Lavoro(**d) for d in docs]


@api_router.post("/lavori", response_model=Lavoro)
async def create_lavoro(payload: LavoroCreate):
    if payload.stato not in ("pianificato", "in_corso", "completato"):
        raise HTTPException(400, "Stato non valido")
    # verifica esistenza cliente
    c = await db.clienti.find_one({"id": payload.cliente_id})
    if not c:
        raise HTTPException(404, "Cliente non trovato")
    lavoro = Lavoro(**{k: v for k, v in payload.model_dump().items() if v is not None})
    await db.lavori.insert_one(serialize(lavoro))
    return lavoro


@api_router.put("/lavori/{lavoro_id}", response_model=Lavoro)
async def update_lavoro(lavoro_id: str, payload: LavoroCreate):
    existing = await db.lavori.find_one({"id": lavoro_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Lavoro non trovato")
    if payload.stato not in ("pianificato", "in_corso", "completato"):
        raise HTTPException(400, "Stato non valido")
    merged = {**existing, **{k: v for k, v in payload.model_dump().items() if v is not None}}
    merged["id"] = lavoro_id
    lavoro = Lavoro(**merged)
    await db.lavori.update_one({"id": lavoro_id}, {"$set": serialize(lavoro)})
    return lavoro


@api_router.delete("/lavori/{lavoro_id}")
async def delete_lavoro(lavoro_id: str):
    res = await db.lavori.delete_one({"id": lavoro_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Lavoro non trovato")
    return {"ok": True}


# ---------- PDF Preventivo ----------

def _euro(v: float) -> str:
    s = f"{v:,.2f}"
    return "€ " + s.replace(",", "X").replace(".", ",").replace("X", ".")


@api_router.get("/clienti/{cliente_id}/preventivo.pdf")
async def preventivo_pdf(cliente_id: str):
    doc = await db.clienti.find_one({"id": cliente_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Cliente non trovato")
    lavori_docs = await db.lavori.find({"cliente_id": cliente_id}, {"_id": 0}).sort("data", -1).to_list(500)
    cantiere_doc = await db.cantiere.find_one({"id": "default"}, {"_id": 0}) or {}
    t_current = await get_tariffe_doc()

    pdf_bytes = _build_preventivo_pdf(doc, lavori_docs, cantiere_doc, t_current)
    filename = f"preventivo_{doc.get('cognome','cliente').lower()}_{doc.get('nome','').lower()}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/export/preventivi.zip")
async def export_tutti_pdf():
    """Esporta un archivio ZIP con un PDF preventivo per ogni cliente."""
    clienti_docs = await db.clienti.find({}, {"_id": 0}).to_list(10000)
    if not clienti_docs:
        raise HTTPException(404, "Nessun cliente da esportare")
    cantiere_doc = await db.cantiere.find_one({"id": "default"}, {"_id": 0}) or {}
    t_current = await get_tariffe_doc()

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for c in clienti_docs:
            lavori_docs = await db.lavori.find({"cliente_id": c["id"]}, {"_id": 0}).sort("data", -1).to_list(500)
            pdf_bytes = _build_preventivo_pdf(c, lavori_docs, cantiere_doc, t_current)
            posto = f"{int(c['posto_barca']):03d}_" if c.get("posto_barca") else ""
            filename = f"{posto}{(c.get('cognome') or 'cliente').lower()}_{(c.get('nome') or '').lower()}.pdf"
            # Sanitize filename
            filename = "".join(ch for ch in filename if ch.isalnum() or ch in "._-")
            zf.writestr(filename, pdf_bytes)
    zip_buf.seek(0)
    zip_filename = f"preventivi_cantiere_{datetime.now().strftime('%Y%m%d')}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={zip_filename}"}
    )


def _build_preventivo_pdf(doc: dict, lavori_docs: list, cantiere_doc: dict, t_current: Tariffe) -> bytes:
    """Genera il PDF preventivo come bytes. Estratto per riuso in singolo + bulk export."""
    buf = io.BytesIO()
    pdf = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=8*mm,
        title=f"Preventivo {doc.get('cognome','')} {doc.get('nome','')}"
    )
    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#0F1B3D")
    TEAK = colors.HexColor("#B0562E")
    SAND = colors.HexColor("#F3EFE7")
    MUTED = colors.HexColor("#5B6478")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=20, textColor=NAVY, spaceAfter=2, leading=22)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=9, textColor=TEAK, spaceBefore=6, spaceAfter=2, leading=11, letterSpace=1.5)
    label = ParagraphStyle("label", parent=styles["Normal"], fontName="Helvetica", fontSize=7, textColor=MUTED, leading=9)
    val = ParagraphStyle("val", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, textColor=NAVY, leading=12)
    body = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica", fontSize=9, textColor=NAVY, leading=11)
    tiny = ParagraphStyle("tiny", parent=styles["Normal"], fontName="Helvetica", fontSize=7, textColor=MUTED, leading=9)

    elems = []

    # Header con logo/nome cantiere + indirizzo
    from reportlab.platypus import Image as RLImage
    import base64 as _b64
    nome_cantiere = (cantiere_doc.get("nome") or "PORTOMARE").upper()
    indirizzo_parts = [x for x in [cantiere_doc.get("indirizzo"), " ".join(filter(None, [cantiere_doc.get("cap"), cantiere_doc.get("citta"), (f"({cantiere_doc.get('provincia')})" if cantiere_doc.get("provincia") else "")])), cantiere_doc.get("telefono"), cantiere_doc.get("email"), cantiere_doc.get("piva") and f"P.IVA {cantiere_doc.get('piva')}"] if x]
    contatti_txt = " · ".join(indirizzo_parts) if indirizzo_parts else ""

    logo_b64 = cantiere_doc.get("logo_base64") or ""
    logo_cell = Paragraph(f"<b>{nome_cantiere}</b>", ParagraphStyle("brand", fontName="Helvetica-Bold", fontSize=18, textColor=NAVY))
    if logo_b64 and "," in logo_b64:
        try:
            raw = _b64.b64decode(logo_b64.split(",", 1)[1])
            logo_cell = RLImage(io.BytesIO(raw), width=30*mm, height=18*mm, kind="proportional")
        except Exception:
            pass

    header_tbl = Table([
        [logo_cell,
         Paragraph(f"<para align=right><font color='#5B6478' size=8>PREVENTIVO</font><br/><font size=14 color='#B0562E'><b>#{doc.get('posto_barca') or '—'}</b></font><br/><font color='#5B6478' size=8>{date.today().strftime('%d/%m/%Y')}</font></para>", body)]
    ], colWidths=[100*mm, 74*mm])
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elems.append(header_tbl)
    if contatti_txt:
        elems.append(Spacer(1, 1*mm))
        elems.append(Paragraph(f"<font color='#5B6478' size=7>{contatti_txt}</font>", body))
    elems.append(Spacer(1, 2*mm))
    # separator
    sep = Table([[""]], colWidths=[186*mm], rowHeights=[1.5])
    sep.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), TEAK)]))
    elems.append(sep)
    elems.append(Spacer(1, 2*mm))

    elems.append(Paragraph("CLIENTE E IMBARCAZIONE", h2))
    potenza = doc.get('potenza_motore') or 0
    litri_pdf = doc.get('litri_olio_motore') or 0
    sosta_label = ('Al coperto' if doc.get('tipo_sosta')=='dentro'
                   else 'Fuori sede' if doc.get('tipo_sosta')=='fuori_sede'
                   else f"Temporanea · {int(doc.get('giorni_sosta_temporanea') or 0)} giorni" if doc.get('tipo_sosta')=='temporanea'
                   else 'Su piazzale (fuori)')
    info_tbl = Table([
        [Paragraph("Cliente", label), Paragraph("Contatti", label)],
        [Paragraph(f"<b>{doc.get('cognome','')} {doc.get('nome','')}</b>", val),
         Paragraph(f"{doc.get('telefono') or '—'} · {doc.get('email') or '—'}", body)],
        [Paragraph("Imbarcazione", label), Paragraph("Sosta", label)],
        [Paragraph(f"<b>{doc.get('tipo_barca','')}</b> · <font color='#5B6478' size=8>L. {doc.get('lunghezza','')} m · Motore: {int(potenza) if potenza else '—'} HP · Olio: {litri_pdf:g} L</font>", body),
         Paragraph(f"<b>{sosta_label}</b> · <font color='#5B6478' size=8>Posto: #{str(doc.get('posto_barca') or '—').zfill(3) if doc.get('posto_barca') else '—'}</font>", body)],
    ], colWidths=[93*mm, 93*mm])
    info_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 1),
        ("TOPPADDING", (0,0), (-1,-1), 0),
    ]))
    elems.append(info_tbl)

    # Costi
    elems.append(Paragraph("DETTAGLIO COSTI ANNUALI", h2))
    voci = []
    def add(label_txt, key):
        v = float(doc.get(key) or 0)
        if v > 0:
            voci.append([label_txt, _euro(v)])
    add("Sosta", "costo_sosta")
    add("Movimentazione", "costo_movimentazione")
    add("Taccaggio", "costo_taccaggio")
    add("Copertura", "costo_copertura")
    # Alaggio/Varo: mostra destinazione se altra
    dest = doc.get("destinazione_alaggio_varo") or "marina_di_campo"
    dest_nome = (doc.get("destinazione_altra_nome") or "").strip()
    if dest == "altra" and dest_nome:
        add(f"Alaggio ({dest_nome})", "costo_alaggio")
        add(f"Varo ({dest_nome})", "costo_varo")
    else:
        add("Alaggio", "costo_alaggio")
        add("Varo", "costo_varo")
    add("Antivegetativa", "costo_antivegetativa")
    add("Magg. scafo sporco", "costo_scafo_sporco")
    add("Lavaggio inizio stagione", "costo_lavaggio_inizio")
    add("Lavaggio fine stagione", "costo_lavaggio_fine")
    add("Manutenzione motore", "costo_manutenzione_motore")
    # Lavorazioni extra: mostrate come voce aggregata nella tabella principale
    lav_extra = doc.get("lavorazioni_extra") or []
    tot_extra = round(sum(float((it or {}).get("prezzo") or 0) for it in lav_extra), 2)
    if tot_extra > 0:
        voci.append(["Lavorazioni extra", _euro(tot_extra)])
    totale = sum(float(doc.get(k) or 0) for k in ("costo_sosta","costo_movimentazione","costo_taccaggio","costo_copertura","costo_alaggio","costo_varo","costo_antivegetativa","costo_scafo_sporco","costo_lavaggio_inizio","costo_lavaggio_fine","costo_manutenzione_motore")) + tot_extra

    if not voci:
        voci = [["Nessun costo configurato", "—"]]

    costi_data = [["VOCE", "IMPORTO"]] + voci + [["TOTALE", _euro(totale)]]
    costi_tbl = Table(costi_data, colWidths=[136*mm, 50*mm])
    n = len(voci)
    costi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), NAVY),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 7),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ("FONTNAME", (0,1), (-1,n), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,n), 9),
        ("TEXTCOLOR", (0,1), (-1,n), NAVY),
        ("ROWBACKGROUNDS", (0,1), (-1,n), [colors.white, SAND]),
        ("LINEBELOW", (0,1), (-1,n), 0.3, colors.HexColor("#D9D9D9")),
        # totale row
        ("BACKGROUND", (0,-1), (-1,-1), TEAK),
        ("TEXTCOLOR", (0,-1), (-1,-1), colors.white),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,-1), (-1,-1), 11),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
        ("RIGHTPADDING", (0,0), (-1,-1), 8),
    ]))
    elems.append(costi_tbl)

    # Dettaglio motore (manodopera + ricambi) se presente
    manodopera = float(doc.get("costo_manodopera_motore") or 0)
    ricambi_tot = float(doc.get("costo_ricambi_totale") or 0)
    has_motore_2 = bool(doc.get("secondo_motore"))
    manodopera_2 = float(doc.get("costo_manodopera_motore_2") or 0)
    ricambi_2_tot = float(doc.get("costo_ricambi_motore_2_totale") or 0)

    def _build_motore_table(title_txt, potenza, litri, litri_piede, nc, nt, girante_on, manod, ric_tot):
        rows = [
            ["Manodopera motore", "", _euro(manod)],
        ]
        if girante_on:
            rows.append(["Girante", "1", _euro(t_current.costo_girante)])
        rows.extend([
            ["Olio motore", f"{litri:g} L", _euro(litri * t_current.costo_olio_motore)],
            ["Filtro olio", "1", _euro(t_current.costo_filtro_olio)],
            ["Candele", str(nc), _euro(nc * t_current.costo_candela)],
            ["Termostato", str(nt), _euro(nt * t_current.costo_termostato)],
            ["Olio piede", f"{litri_piede:g} L", _euro(litri_piede * t_current.costo_olio_piede)],
            ["Kit anodi interni", "1", _euro(t_current.costo_anodi_interni)],
            ["Kit anodi esterni", "1", _euro(t_current.costo_anodi_esterni)],
            ["Ingrassaggio", "1", _euro(t_current.costo_ingrassaggio)],
        ])
        subtotale = manod + ric_tot
        header_row = [f"{title_txt} — {int(potenza) if potenza else 0} HP", "", _euro(subtotale)]
        data = [header_row, ["VOCE", "Q.TÀ", "IMPORTO"]] + rows
        tbl = Table(data, colWidths=[116*mm, 20*mm, 50*mm])
        tbl.setStyle(TableStyle([
            # Titolo motore
            ("SPAN", (0,0), (1,0)),
            ("BACKGROUND", (0,0), (-1,0), TEAK),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("ALIGN", (2,0), (2,0), "RIGHT"),
            # Intestazione tabella
            ("BACKGROUND", (0,1), (-1,1), NAVY),
            ("TEXTCOLOR", (0,1), (-1,1), colors.white),
            ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"),
            ("FONTSIZE", (0,1), (-1,1), 7),
            ("ALIGN", (1,1), (2,-1), "RIGHT"),
            # Righe
            ("FONTNAME", (0,2), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,2), (-1,-1), 8),
            ("TEXTCOLOR", (0,2), (-1,-1), NAVY),
            ("ROWBACKGROUNDS", (0,2), (-1,-1), [colors.white, SAND]),
            ("LINEBELOW", (0,2), (-1,-1), 0.3, colors.HexColor("#D9D9D9")),
            ("TOPPADDING", (0,0), (-1,-1), 1),
            ("BOTTOMPADDING", (0,0), (-1,-1), 1),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ]))
        return tbl

    if manodopera > 0 or ricambi_tot > 0 or manodopera_2 > 0 or ricambi_2_tot > 0:
        elems.append(Paragraph("DETTAGLIO MANUTENZIONE MOTORE", h2))
        # 1° motore
        nc = int(doc.get("numero_candele") or 0)
        nt = int(doc.get("numero_termostati") or 0)
        girante_attivo = bool(doc.get("girante_attivo", True))
        litri = float(doc.get("litri_olio_motore") or 0)
        potenza_1 = float(doc.get("potenza_motore") or 0)
        motore_1_label = "1° Motore" if has_motore_2 else "Motore"
        litri_piede_1 = float(doc.get("litri_olio_piede") or 1.0)
        elems.append(_build_motore_table(motore_1_label, potenza_1, litri, litri_piede_1, nc, nt, girante_attivo, manodopera, ricambi_tot))

        # 2° motore
        if has_motore_2:
            elems.append(Spacer(1, 1*mm))
            nc2 = int(doc.get("numero_candele_2") or 0)
            nt2 = int(doc.get("numero_termostati_2") or 0)
            girante_2 = bool(doc.get("girante_2_attivo", True))
            litri2 = float(doc.get("litri_olio_motore_2") or 0)
            litri_piede_2 = float(doc.get("litri_olio_piede_2") or 1.0)
            potenza_2 = float(doc.get("potenza_motore_2") or 0)
            elems.append(_build_motore_table("2° Motore", potenza_2, litri2, litri_piede_2, nc2, nt2, girante_2, manodopera_2, ricambi_2_tot))

    # Lavorazioni extra dettaglio
    if lav_extra and tot_extra > 0:
        elems.append(Paragraph("LAVORAZIONI EXTRA", h2))
        rows_extra = [["DESCRIZIONE", "IMPORTO"]]
        for it in lav_extra:
            desc = (it.get("descrizione") or "").strip() or "—"
            prezzo = float(it.get("prezzo") or 0)
            if prezzo > 0 or desc != "—":
                rows_extra.append([desc[:80], _euro(prezzo)])
        rows_extra.append(["TOTALE EXTRA", _euro(tot_extra)])
        ex_tbl = Table(rows_extra, colWidths=[136*mm, 50*mm])
        n_ex = len(rows_extra) - 2
        ex_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), NAVY),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 7),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
            ("FONTNAME", (0,1), (-1,n_ex), "Helvetica"),
            ("FONTSIZE", (0,1), (-1,n_ex), 9),
            ("TEXTCOLOR", (0,1), (-1,n_ex), NAVY),
            ("ROWBACKGROUNDS", (0,1), (-1,n_ex), [colors.white, SAND]),
            ("LINEBELOW", (0,1), (-1,n_ex), 0.3, colors.HexColor("#D9D9D9")),
            ("BACKGROUND", (0,-1), (-1,-1), TEAK),
            ("TEXTCOLOR", (0,-1), (-1,-1), colors.white),
            ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,-1), (-1,-1), 10),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ]))
        elems.append(ex_tbl)

    # Scadenze
    if doc.get("scadenza_antivegetativa") or doc.get("scadenza_manutenzione"):
        elems.append(Paragraph("PROSSIME SCADENZE", h2))
        rows = []
        if doc.get("scadenza_antivegetativa"):
            rows.append(["Antivegetativa", doc["scadenza_antivegetativa"]])
        if doc.get("scadenza_manutenzione"):
            rows.append(["Manutenzione motore", doc["scadenza_manutenzione"]])
        sc_tbl = Table(rows, colWidths=[136*mm, 50*mm])
        sc_tbl.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("TEXTCOLOR", (0,0), (-1,-1), NAVY),
            ("ALIGN", (1,0), (1,-1), "RIGHT"),
            ("LINEBELOW", (0,0), (-1,-1), 0.3, colors.HexColor("#D9D9D9")),
            ("TOPPADDING", (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ]))
        elems.append(sc_tbl)

    # Storico lavori strutturato (limitato a 5 righe per garantire 1 pagina)
    if lavori_docs:
        elems.append(Paragraph("STORICO LAVORI ESEGUITI", h2))
        headers = ["Data", "Tipo", "Descrizione", "Costo"]
        rows = [headers]
        for l in lavori_docs[:5]:
            rows.append([
                l.get("data",""),
                l.get("tipo",""),
                (l.get("descrizione","") or "")[:60],
                _euro(float(l.get("costo") or 0)),
            ])
        lav_tbl = Table(rows, colWidths=[22*mm, 38*mm, 94*mm, 32*mm])
        lav_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), NAVY),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 7),
            ("FONTSIZE", (0,1), (-1,-1), 8),
            ("TEXTCOLOR", (0,1), (-1,-1), NAVY),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, SAND]),
            ("ALIGN", (3,0), (3,-1), "RIGHT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ]))
        elems.append(lav_tbl)

    # Note
    if doc.get("note_lavori"):
        elems.append(Paragraph("NOTE", h2))
        elems.append(Paragraph(doc["note_lavori"].replace("\n", "<br/>"), body))

    elems.append(Spacer(1, 3*mm))
    footer_name = cantiere_doc.get("nome") or "Portomare"
    elems.append(Paragraph(
        f"Documento generato automaticamente da {footer_name} — Gestione Cantiere Nautico. "
        f"Validità 30 giorni dalla data di emissione ({date.today().strftime('%d/%m/%Y')}).",
        tiny
    ))

    pdf.build(elems)
    buf.seek(0)
    return buf.getvalue()


# ---------- CANTIERE INFO (logo + indirizzo) ----------

class Cantiere(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: "default")
    nome: str = "Portomare"
    slogan: str = "Cantiere nautico dal 1985"
    indirizzo: str = ""
    citta: str = ""
    cap: str = ""
    provincia: str = ""
    telefono: str = ""
    email: str = ""
    piva: str = ""
    sito_web: str = ""
    orari: str = ""
    logo_base64: str = ""  # data:image/...;base64,...
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class CantiereUpdate(BaseModel):
    nome: Optional[str] = None
    slogan: Optional[str] = None
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    telefono: Optional[str] = None
    email: Optional[str] = None
    piva: Optional[str] = None
    sito_web: Optional[str] = None
    orari: Optional[str] = None
    logo_base64: Optional[str] = None


@api_router.get("/cantiere", response_model=Cantiere)
async def get_cantiere():
    doc = await db.cantiere.find_one({"id": "default"}, {"_id": 0})
    if not doc:
        c = Cantiere()
        await db.cantiere.insert_one(serialize(c))
        return c
    return Cantiere(**doc)


@api_router.put("/cantiere", response_model=Cantiere)
async def update_cantiere(payload: CantiereUpdate):
    current = await get_cantiere()
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    new_data = current.model_dump()
    new_data.update(updates)
    new_data["updated_at"] = datetime.now(timezone.utc)
    c = Cantiere(**new_data)
    await db.cantiere.update_one({"id": "default"}, {"$set": serialize(c)}, upsert=True)
    return c


# ---------- BACKUP / RESTORE ----------

@api_router.get("/backup")
async def backup_data():
    """Esporta tutti i dati del cantiere in un unico JSON scaricabile."""
    import json as _json
    clienti = await db.clienti.find({}, {"_id": 0}).to_list(10000)
    lavori = await db.lavori.find({}, {"_id": 0}).to_list(10000)
    tariffe = await db.tariffe.find_one({"id": "default"}, {"_id": 0})
    cantiere = await db.cantiere.find_one({"id": "default"}, {"_id": 0})

    payload = {
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "app": "Portomare Cantiere Nautico",
        "cantiere": cantiere,
        "tariffe": tariffe,
        "clienti": clienti,
        "lavori": lavori,
        "counts": {
            "clienti": len(clienti),
            "lavori": len(lavori),
        }
    }
    body = _json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    filename = f"backup_cantiere_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(
        iter([body]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class PreventivoInline(ClienteCreate):
    """Payload preventivo veloce: solo nome e cognome sono obbligatori.
    Tutti gli altri campi hanno default sensati per generare un PDF anche con dati minimi.
    """
    nome: str
    cognome: str
    tipo_barca: Optional[str] = "—"
    lunghezza: Optional[float] = 0.0
    tipo_sosta: Optional[str] = "dentro"


@api_router.post("/preventivo/pdf")
async def preventivo_pdf_inline(payload: PreventivoInline):
    """Genera un PDF preventivo al volo senza salvare in DB. Basta nome+cognome."""
    if not payload.nome.strip() or not payload.cognome.strip():
        raise HTTPException(400, "Nome e cognome sono obbligatori")
    if payload.tipo_sosta not in ("dentro", "fuori", "fuori_sede", "temporanea"):
        payload.tipo_sosta = "dentro"

    t = await get_tariffe_doc()
    cantiere_doc = await db.cantiere.find_one({"id": "default"}, {"_id": 0}) or {}

    lunghezza = float(payload.lunghezza or 0)
    auto_costi = calcola_costi(
        lunghezza, payload.tipo_sosta or "dentro", t,
        float(payload.potenza_motore or 0),
        int(payload.numero_candele or 4),
        int(payload.numero_termostati or 1),
        bool(payload.antivegetativa_attiva if payload.antivegetativa_attiva is not None else True),
        bool(payload.girante_attivo if payload.girante_attivo is not None else True),
        float(payload.litri_olio_motore if payload.litri_olio_motore is not None else 3.0),
        bool(payload.lavaggio_inizio_attivo if payload.lavaggio_inizio_attivo is not None else True),
        bool(payload.lavaggio_fine_attivo if payload.lavaggio_fine_attivo is not None else True),
        bool(payload.secondo_motore if payload.secondo_motore is not None else False),
        float(payload.potenza_motore_2 or 0),
        float(payload.litri_olio_motore_2 if payload.litri_olio_motore_2 is not None else 3.0),
        int(payload.numero_candele_2 or 4),
        int(payload.numero_termostati_2 or 1),
        bool(payload.girante_2_attivo if payload.girante_2_attivo is not None else True),
        bool(payload.scafo_sporco_attivo if payload.scafo_sporco_attivo is not None else False),
        bool(payload.copertura_attiva if payload.copertura_attiva is not None else False),
        float(payload.litri_olio_piede if payload.litri_olio_piede is not None else 1.0),
        float(payload.litri_olio_piede_2 if payload.litri_olio_piede_2 is not None else 1.0),
        int(payload.giorni_sosta_temporanea if payload.giorni_sosta_temporanea is not None else 0),
        (payload.destinazione_alaggio_varo or "marina_di_campo"),
        bool(payload.alaggio_varo_attivo if payload.alaggio_varo_attivo is not None else False),
    )
    auto_costi.pop("ricambi_dettaglio", None)
    auto_costi.pop("ricambi_2_dettaglio", None)

    doc = payload.model_dump()
    manual_alaggio_varo = (payload.alaggio_varo_attivo and payload.destinazione_alaggio_varo == "altra")
    # Applica costi calcolati (auto); alaggio/varo manuali se destinazione="altra"
    for k, v in auto_costi.items():
        if manual_alaggio_varo and k in ("costo_alaggio", "costo_varo"):
            existing_val = doc.get(k)
            doc[k] = float(existing_val) if existing_val is not None else 0.0
        else:
            doc[k] = v
    # Normalize lavorazioni_extra
    doc["lavorazioni_extra"] = _sanitize_lavorazioni_extra(doc.get("lavorazioni_extra"))

    pdf_bytes = _build_preventivo_pdf(doc, [], cantiere_doc, t)
    filename = f"preventivo_{payload.cognome.strip()}_{payload.nome.strip()}.pdf".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class RestoreRequest(BaseModel):
    version: Optional[int] = None
    cantiere: Optional[dict] = None
    tariffe: Optional[dict] = None
    clienti: Optional[List[dict]] = None
    lavori: Optional[List[dict]] = None


@api_router.post("/restore")
async def restore_data(payload: RestoreRequest):
    """Ripristina i dati dal backup JSON. Sovrascrive completamente il DB."""
    restored = {"clienti": 0, "lavori": 0, "tariffe": False, "cantiere": False}

    if payload.cantiere is not None:
        c = Cantiere(**{k: v for k, v in payload.cantiere.items() if k in Cantiere.model_fields})
        await db.cantiere.delete_many({})
        await db.cantiere.insert_one(serialize(c))
        restored["cantiere"] = True

    if payload.tariffe is not None:
        t = Tariffe(**{k: v for k, v in payload.tariffe.items() if k in Tariffe.model_fields})
        await db.tariffe.delete_many({})
        await db.tariffe.insert_one(serialize(t))
        restored["tariffe"] = True

    if payload.clienti is not None:
        await db.clienti.delete_many({})
        docs = []
        for c in payload.clienti:
            try:
                cli = Cliente(**{k: v for k, v in c.items() if k in Cliente.model_fields})
                docs.append(serialize(cli))
            except Exception:
                pass
        if docs:
            await db.clienti.insert_many(docs)
        restored["clienti"] = len(docs)

    if payload.lavori is not None:
        await db.lavori.delete_many({})
        docs = []
        for l in payload.lavori:
            try:
                lav = Lavoro(**{k: v for k, v in l.items() if k in Lavoro.model_fields})
                docs.append(serialize(lav))
            except Exception:
                pass
        if docs:
            await db.lavori.insert_many(docs)
        restored["lavori"] = len(docs)

    return {"ok": True, "restored": restored}


# ---------- AUTH ROUTES ----------

class LoginRequest(BaseModel):
    email: str
    password: str


class RegisterRequest(BaseModel):
    email: str
    password: str
    nome: Optional[str] = ""


auth_router = APIRouter(prefix="/api/auth")


def _set_cookie(response: Response, token: str):
    response.set_cookie(
        key="access_token", value=token, httponly=True, secure=False,
        samesite="lax", max_age=8 * 3600, path="/",
    )


@auth_router.post("/register")
async def register(payload: RegisterRequest, response: Response):
    email = payload.email.strip().lower()
    if not email or not payload.password:
        raise HTTPException(400, "Email e password obbligatorie")
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Email già registrata")
    user_id = str(uuid.uuid4())
    doc = {
        "id": user_id,
        "email": email,
        "password_hash": hash_password(payload.password),
        "nome": payload.nome or email.split("@")[0],
        "role": "user",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    token = create_access_token(user_id, email)
    _set_cookie(response, token)
    return {"id": user_id, "email": email, "nome": doc["nome"], "role": "user", "token": token}


@auth_router.post("/login")
async def login(payload: LoginRequest, response: Response):
    email = payload.email.strip().lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(401, "Email o password non corretta")
    token = create_access_token(user["id"], email)
    _set_cookie(response, token)
    return {"id": user["id"], "email": email, "nome": user.get("nome", ""), "role": user.get("role", "user"), "token": token}


@auth_router.post("/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@auth_router.get("/me")
async def me(user: dict = Depends(get_current_user)):
    return {"id": user["id"], "email": user["email"], "nome": user.get("nome", ""), "role": user.get("role", "user")}


app.include_router(auth_router)


async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@portomare.it").strip().lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "portomare2026")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "nome": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Admin seeded: {admin_email}")
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info(f"Admin password updated: {admin_email}")


@app.on_event("startup")
async def _startup():
    await db.users.create_index("email", unique=True)
    await seed_admin()
    # Migrazione iter13: chi aveva scafo_sporco applicato (costo > 0) senza il nuovo flag scafo_sporco_attivo
    # imposta il flag a True per preservare il comportamento pre-esistente. Idempotente.
    try:
        await db.clienti.update_many(
            {"scafo_sporco_attivo": {"$exists": False}, "costo_scafo_sporco": {"$gt": 0}},
            {"$set": {"scafo_sporco_attivo": True}},
        )
        await db.clienti.update_many(
            {"scafo_sporco_attivo": {"$exists": False}},
            {"$set": {"scafo_sporco_attivo": False}},
        )
    except Exception as e:
        logger.warning(f"Migration iter13 scafo_sporco_attivo skipped: {e}")
    # Migrazione iter14: chi aveva costo_copertura > 0 → copertura_attiva=True
    try:
        await db.clienti.update_many(
            {"copertura_attiva": {"$exists": False}, "costo_copertura": {"$gt": 0}},
            {"$set": {"copertura_attiva": True}},
        )
        await db.clienti.update_many(
            {"copertura_attiva": {"$exists": False}},
            {"$set": {"copertura_attiva": False}},
        )
    except Exception as e:
        logger.warning(f"Migration iter14 copertura_attiva skipped: {e}")
    # Migrazione iter20: chi aveva tipo_sosta="fuori" → alaggio_varo_attivo=True (preserva comportamento pregresso)
    try:
        await db.clienti.update_many(
            {"alaggio_varo_attivo": {"$exists": False}, "tipo_sosta": "fuori"},
            {"$set": {"alaggio_varo_attivo": True}},
        )
        await db.clienti.update_many(
            {"alaggio_varo_attivo": {"$exists": False}},
            {"$set": {"alaggio_varo_attivo": False}},
        )
    except Exception as e:
        logger.warning(f"Migration iter20 alaggio_varo_attivo skipped: {e}")


# ---------- REPORT INCASSI ----------

@api_router.get("/report/incassi")
async def report_incassi(anno: Optional[int] = None):
    """Sommatorie per categoria su tutti i clienti (filtrabili per anno)."""
    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(10000)

    def s(key):
        return round(sum(float(d.get(key) or 0) for d in docs), 2)

    incasso_sosta = s("costo_sosta")
    incasso_movimentazione = s("costo_movimentazione")
    incasso_taccaggio = s("costo_taccaggio")
    incasso_alaggio = s("costo_alaggio")
    incasso_varo = s("costo_varo")
    incasso_coperture = s("costo_copertura")
    incasso_antivegetativa = s("costo_antivegetativa")
    incasso_scafo_sporco = s("costo_scafo_sporco")
    incasso_lavaggio_inizio = s("costo_lavaggio_inizio")
    incasso_lavaggio_fine = s("costo_lavaggio_fine")
    incasso_motore = s("costo_manutenzione_motore")
    incasso_lavorazioni_extra = round(sum(_totale_extra(d) for d in docs), 2)

    # Suddivisione motore
    incasso_manodopera = s("costo_manodopera_motore")
    incasso_ricambi = s("costo_ricambi_totale")

    totale = round(
        incasso_sosta + incasso_movimentazione + incasso_taccaggio +
        incasso_alaggio + incasso_varo + incasso_coperture +
        incasso_antivegetativa + incasso_scafo_sporco +
        incasso_lavaggio_inizio + incasso_lavaggio_fine +
        incasso_motore + incasso_lavorazioni_extra, 2
    )

    # Ripartizione per tipo sosta
    per_tipo_sosta = {"dentro": 0.0, "fuori": 0.0, "fuori_sede": 0.0, "temporanea": 0.0}
    for d in docs:
        tipo = d.get("tipo_sosta")
        if tipo in per_tipo_sosta:
            client_tot = sum(float(d.get(k) or 0) for k in (
                "costo_sosta","costo_movimentazione","costo_taccaggio",
                "costo_alaggio","costo_varo","costo_copertura",
                "costo_antivegetativa","costo_scafo_sporco",
                "costo_lavaggio_inizio","costo_lavaggio_fine",
                "costo_manutenzione_motore"
            )) + _totale_extra(d)
            per_tipo_sosta[tipo] = round(per_tipo_sosta[tipo] + client_tot, 2)

    return {
        "totale_clienti": len(docs),
        "totale": totale,
        "categorie": {
            "sosta": incasso_sosta,
            "movimentazione_taccaggio": round(incasso_movimentazione + incasso_taccaggio, 2),
            "alaggio_varo": round(incasso_alaggio + incasso_varo, 2),
            "coperture": incasso_coperture,
            "antivegetativa": incasso_antivegetativa,
            "scafo_sporco": incasso_scafo_sporco,
            "lavaggi": round(incasso_lavaggio_inizio + incasso_lavaggio_fine, 2),
            "manutenzione_motore": incasso_motore,
            "lavorazioni_extra": incasso_lavorazioni_extra,
        },
        "motore_dettaglio": {
            "manodopera": incasso_manodopera,
            "ricambi": incasso_ricambi,
        },
        "sosta_dettaglio": {
            "sosta": incasso_sosta,
            "movimentazione": incasso_movimentazione,
            "taccaggio": incasso_taccaggio,
        },
        "alaggio_varo_dettaglio": {
            "alaggio": incasso_alaggio,
            "varo": incasso_varo,
        },
        "lavaggi_dettaglio": {
            "inizio_stagione": incasso_lavaggio_inizio,
            "fine_stagione": incasso_lavaggio_fine,
        },
        "per_tipo_sosta": per_tipo_sosta,
    }


# ---------- GESTIONE ANNI ----------

class PagatoUpdate(BaseModel):
    pagato: bool


@api_router.patch("/clienti/{cliente_id}/pagato")
async def toggle_pagato(cliente_id: str, payload: PagatoUpdate):
    """Aggiorna lo stato di pagamento del cliente."""
    existing = await db.clienti.find_one({"id": cliente_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Cliente non trovato")
    update = {
        "pagato": bool(payload.pagato),
        "data_pagamento": datetime.now(timezone.utc).date().isoformat() if payload.pagato else None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.clienti.update_one({"id": cliente_id}, {"$set": update})
    return {"ok": True, "cliente_id": cliente_id, **update}


@api_router.get("/report/pagamenti")
async def report_pagamenti(anno: Optional[int] = None):
    """Elenco clienti con stato pagamento e totale dovuto (per anno)."""
    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(10000)
    docs.sort(key=lambda d: ((d.get("cognome") or "").strip().lower(), (d.get("nome") or "").strip().lower()))

    result = []
    for d in docs:
        totale = sum(float(d.get(k) or 0) for k in (
            "costo_sosta","costo_movimentazione","costo_taccaggio",
            "costo_copertura","costo_alaggio","costo_varo",
            "costo_antivegetativa","costo_scafo_sporco",
            "costo_lavaggio_inizio","costo_lavaggio_fine",
            "costo_manutenzione_motore"
        )) + _totale_extra(d)
        result.append({
            "id": d["id"],
            "nome": d.get("nome",""),
            "cognome": d.get("cognome",""),
            "tipo_barca": d.get("tipo_barca",""),
            "posto_barca": d.get("posto_barca"),
            "tipo_sosta": d.get("tipo_sosta"),
            "totale": round(totale, 2),
            "pagato": bool(d.get("pagato", False)),
            "data_pagamento": d.get("data_pagamento"),
        })

    totale_pagato = sum(c["totale"] for c in result if c["pagato"])
    totale_da_pagare = sum(c["totale"] for c in result if not c["pagato"])
    return {
        "clienti": result,
        "totale_pagato": round(totale_pagato, 2),
        "totale_da_pagare": round(totale_da_pagare, 2),
        "numero_pagati": sum(1 for c in result if c["pagato"]),
        "numero_non_pagati": sum(1 for c in result if not c["pagato"]),
    }


@api_router.get("/report/pagamenti.pdf")
async def report_pagamenti_pdf(anno: Optional[int] = None, stato: str = "tutti"):
    """Genera PDF stampabile del report pagamenti. stato: tutti|pagati|non_pagati"""
    if stato not in ("tutti", "pagati", "non_pagati"):
        raise HTTPException(400, "stato deve essere 'tutti', 'pagati' o 'non_pagati'")

    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(10000)
    docs.sort(key=lambda d: ((d.get("cognome") or "").strip().lower(), (d.get("nome") or "").strip().lower()))
    cantiere_doc = await db.cantiere.find_one({"id": "default"}, {"_id": 0}) or {}

    rows_all = []
    for d in docs:
        totale = sum(float(d.get(k) or 0) for k in (
            "costo_sosta","costo_movimentazione","costo_taccaggio",
            "costo_copertura","costo_alaggio","costo_varo",
            "costo_antivegetativa","costo_scafo_sporco",
            "costo_lavaggio_inizio","costo_lavaggio_fine",
            "costo_manutenzione_motore"
        )) + _totale_extra(d)
        rows_all.append({
            "cognome": d.get("cognome", ""),
            "nome": d.get("nome", ""),
            "tipo_barca": d.get("tipo_barca", ""),
            "posto_barca": d.get("posto_barca"),
            "totale": round(totale, 2),
            "pagato": bool(d.get("pagato", False)),
            "data_pagamento": d.get("data_pagamento"),
        })

    if stato == "pagati":
        rows = [r for r in rows_all if r["pagato"]]
    elif stato == "non_pagati":
        rows = [r for r in rows_all if not r["pagato"]]
    else:
        rows = rows_all

    tot_pagati = sum(r["totale"] for r in rows_all if r["pagato"])
    tot_non_pagati = sum(r["totale"] for r in rows_all if not r["pagato"])

    # Build PDF
    buf = io.BytesIO()
    pdf = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm,
        title=f"Report pagamenti {anno or ''}"
    )
    styles = getSampleStyleSheet()
    NAVY = colors.HexColor("#0F1B3D")
    TEAK = colors.HexColor("#B0562E")
    SAND = colors.HexColor("#F3EFE7")
    GREEN = colors.HexColor("#16803C")
    RED = colors.HexColor("#B91C1C")
    MUTED = colors.HexColor("#5B6478")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=20, textColor=NAVY, spaceAfter=4, leading=24)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=10, textColor=TEAK, spaceBefore=10, spaceAfter=6, leading=12)
    body = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica", fontSize=9, textColor=NAVY, leading=12)
    tiny = ParagraphStyle("tiny", parent=styles["Normal"], fontName="Helvetica", fontSize=8, textColor=MUTED, leading=10)

    elems = []
    nome_cantiere = (cantiere_doc.get("nome") or "PORTOMARE").upper()
    stato_label = {"tutti": "Tutti", "pagati": "Solo pagati", "non_pagati": "Solo non pagati"}[stato]
    anno_label = str(anno) if anno else "Tutti gli anni"

    # Header
    elems.append(Paragraph(f"<b>{nome_cantiere}</b>", h1))
    elems.append(Paragraph(f"<font color='#5B6478' size=9>REPORT PAGAMENTI · Anno {anno_label} · Filtro: {stato_label} · Emesso: {date.today().strftime('%d/%m/%Y')}</font>", body))
    sep = Table([[""]], colWidths=[180*mm], rowHeights=[2])
    sep.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), TEAK)]))
    elems.append(Spacer(1, 3*mm))
    elems.append(sep)
    elems.append(Spacer(1, 5*mm))

    # Riepilogo
    riepilogo = Table([
        [
            Paragraph(f"<b>Clienti totali</b><br/><font size=14>{len(rows_all)}</font>", body),
            Paragraph(f"<b><font color='#16803C'>Pagati</font></b><br/><font size=14 color='#16803C'>{sum(1 for r in rows_all if r['pagato'])} · {_euro(tot_pagati)}</font>", body),
            Paragraph(f"<b><font color='#B91C1C'>Non pagati</font></b><br/><font size=14 color='#B91C1C'>{sum(1 for r in rows_all if not r['pagato'])} · {_euro(tot_non_pagati)}</font>", body),
        ]
    ], colWidths=[60*mm, 60*mm, 60*mm])
    riepilogo.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ("INNERGRID", (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
    ]))
    elems.append(riepilogo)
    elems.append(Spacer(1, 6*mm))

    # Tabella clienti
    header = ["Posto", "Cliente", "Barca", "Totale", "Stato", "Data pag."]
    table_data = [header]
    for r in rows:
        stato_cell = "PAGATO" if r["pagato"] else "NON PAGATO"
        table_data.append([
            f"#{int(r['posto_barca']):03d}" if r["posto_barca"] else "—",
            f"{r['cognome']} {r['nome']}".strip(),
            (r["tipo_barca"] or "")[:30],
            _euro(r["totale"]),
            stato_cell,
            r["data_pagamento"] or "—",
        ])
    # Totale finale (solo delle righe filtrate)
    tot_filtered = sum(r["totale"] for r in rows)
    table_data.append(["", "", "TOTALE", _euro(tot_filtered), "", ""])

    tbl = Table(table_data, colWidths=[18*mm, 52*mm, 40*mm, 28*mm, 26*mm, 22*mm], repeatRows=1)
    n = len(rows)
    style = [
        ("BACKGROUND", (0,0), (-1,0), NAVY),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 8),
        ("ALIGN", (3,0), (3,-1), "RIGHT"),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
        ("ALIGN", (4,0), (4,-1), "CENTER"),
        ("FONTNAME", (0,1), (-1,n), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,n), 8),
        ("TEXTCOLOR", (0,1), (-1,n), NAVY),
        ("ROWBACKGROUNDS", (0,1), (-1,n), [colors.white, SAND]),
        ("LINEBELOW", (0,1), (-1,n), 0.3, colors.HexColor("#D9D9D9")),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        # Riga totale
        ("BACKGROUND", (0,-1), (-1,-1), TEAK),
        ("TEXTCOLOR", (0,-1), (-1,-1), colors.white),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,-1), (-1,-1), 10),
        ("SPAN", (0,-1), (2,-1)),
    ]
    # Colora colonna stato per ogni riga in base a pagato
    for i, r in enumerate(rows, start=1):
        col = GREEN if r["pagato"] else RED
        style.append(("TEXTCOLOR", (4,i), (4,i), col))
        style.append(("FONTNAME", (4,i), (4,i), "Helvetica-Bold"))
    tbl.setStyle(TableStyle(style))
    elems.append(tbl)

    if not rows:
        elems.append(Spacer(1, 4*mm))
        elems.append(Paragraph("<i>Nessun cliente corrisponde al filtro selezionato.</i>", body))

    elems.append(Spacer(1, 8*mm))
    elems.append(Paragraph(
        f"Documento generato automaticamente da {cantiere_doc.get('nome') or 'Portomare'} — {date.today().strftime('%d/%m/%Y %H:%M')}. "
        f"Il totale filtrato include solo i clienti visibili.",
        tiny
    ))

    pdf.build(elems)
    buf.seek(0)
    filename = f"report_pagamenti_{anno_label.replace(' ', '_').lower()}_{stato}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/anni")
async def list_anni():
    """Ritorna la lista degli anni con conteggio clienti per anno."""
    now_year = datetime.now().year
    docs = await db.clienti.find({}, {"anno": 1, "_id": 0}).to_list(20000)
    counts = {}
    for d in docs:
        y = d.get("anno") or now_year
        counts[y] = counts.get(y, 0) + 1
    if now_year not in counts:
        counts[now_year] = counts.get(now_year, 0)

    anni_sorted = sorted(counts.keys(), reverse=True)
    return {
        "anno_corrente": now_year,
        "anni": [{"anno": y, "clienti": counts[y]} for y in anni_sorted],
    }


class ApriAnnoRequest(BaseModel):
    anno: int
    duplica_da: Optional[int] = None  # anno da cui duplicare i clienti


@api_router.post("/anni/apri")
async def apri_anno(payload: ApriAnnoRequest):
    """Apre un nuovo anno. Se duplica_da è specificato, copia i clienti da quell'anno (ricalcolando i costi con le tariffe correnti)."""
    if payload.anno < 2000 or payload.anno > 2100:
        raise HTTPException(400, "Anno non valido")

    # Verifica se ci sono già clienti per quest'anno
    existing_count = await db.clienti.count_documents({"anno": payload.anno})

    duplicati = 0
    if payload.duplica_da is not None and existing_count == 0:
        origine = await db.clienti.find({"anno": payload.duplica_da}, {"_id": 0}).to_list(10000)
        t = await get_tariffe_doc()
        for c in origine:
            # Nuovo ID e anno, ricalcola costi
            new_id = str(uuid.uuid4())
            auto_costi = calcola_costi(
                c.get("lunghezza", 0), c.get("tipo_sosta", "dentro"), t,
                c.get("potenza_motore", 0) or 0,
                c.get("numero_candele", 4) or 4,
                c.get("numero_termostati", 1) or 1,
                bool(c.get("antivegetativa_attiva", True)),
                bool(c.get("girante_attivo", True)),
                float(c.get("litri_olio_motore", 3.0) or 3.0),
                bool(c.get("lavaggio_inizio_attivo", True)),
                bool(c.get("lavaggio_fine_attivo", True)),
                bool(c.get("secondo_motore", False)),
                float(c.get("potenza_motore_2", 0) or 0),
                float(c.get("litri_olio_motore_2", 3.0) or 3.0),
                int(c.get("numero_candele_2", 4) or 4),
                int(c.get("numero_termostati_2", 1) or 1),
                bool(c.get("girante_2_attivo", True)),
                bool(c.get("scafo_sporco_attivo", False)),
                bool(c.get("copertura_attiva", False)),
                float(c.get("litri_olio_piede", 1.0) or 1.0),
                float(c.get("litri_olio_piede_2", 1.0) or 1.0),
                int(c.get("giorni_sosta_temporanea", 0) or 0),
                str(c.get("destinazione_alaggio_varo") or "marina_di_campo"),
                bool(c.get("alaggio_varo_attivo", False)),
            )
            auto_costi.pop("ricambi_dettaglio", None)
            auto_costi.pop("ricambi_2_dettaglio", None)

            data = {**c, **auto_costi, "id": new_id, "anno": payload.anno,
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc),
                    # Reset note e scadenze del nuovo anno
                    "note_lavori": "", "scadenza_antivegetativa": None, "scadenza_manutenzione": None}
            try:
                cli = Cliente(**{k: v for k, v in data.items() if k in Cliente.model_fields or k in ("posto_barca", "scadenza_antivegetativa", "scadenza_manutenzione")})
                await db.clienti.insert_one(serialize(cli))
                duplicati += 1
            except Exception as e:
                logger.warning(f"Errore duplicazione cliente: {e}")

    return {"ok": True, "anno": payload.anno, "duplicati": duplicati, "gia_esistenti": existing_count}


@api_router.delete("/anni/{anno}")
async def elimina_anno(anno: int):
    """Elimina tutti i clienti e lavori di un anno specifico."""
    if anno == datetime.now().year:
        # Permettiamo di svuotare anche l'anno corrente ma con conteggio
        pass
    res_clienti = await db.clienti.delete_many({"anno": anno})
    # Recupera ID clienti eliminati per pulire lavori? Meglio filtrare per anno lavoro
    res_lavori = await db.lavori.delete_many({"anno": anno})
    return {
        "ok": True,
        "anno": anno,
        "clienti_eliminati": res_clienti.deleted_count,
        "lavori_eliminati": res_lavori.deleted_count,
    }


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

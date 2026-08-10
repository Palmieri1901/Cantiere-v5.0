"""Export CSV, Excel (commercialista), ZIP di tutti i preventivi PDF."""
import io
import zipfile
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import pandas as pd

from database import db
from helpers import get_tariffe_doc
from pdf_builders import _build_preventivo_pdf

router = APIRouter()


@router.get("/export/clienti.csv")
async def export_csv():
    docs = await db.clienti.find({}, {"_id": 0}).to_list(1000)
    if not docs:
        docs = [{}]
    df = pd.DataFrame(docs)
    if not df.empty:
        cols_order = ["posto_barca", "nome", "cognome", "tipo_barca", "lunghezza", "tipo_sosta",
                      "telefono", "email", "costo_sosta", "costo_copertura", "costo_alaggio",
                      "costo_varo", "costo_antivegetativa", "costo_manutenzione_motore",
                      "scadenza_antivegetativa", "scadenza_manutenzione", "note_lavori"]
        cols = [c for c in cols_order if c in df.columns] + [c for c in df.columns if c not in cols_order]
        df = df[cols]

    buf = io.StringIO()
    df.to_csv(buf, index=False, sep=";")
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clienti_cantiere.csv"}
    )


@router.get("/export/clienti.xlsx")
async def export_xlsx(anno: Optional[int] = None):
    q = {}
    if anno is not None:
        q["anno"] = anno
    docs = await db.clienti.find(q, {"_id": 0}).to_list(1000)
    docs.sort(key=lambda d: ((d.get("cognome") or "").strip().lower(), (d.get("nome") or "").strip().lower()))

    COLS = [
        ("Anno", "anno"),
        ("Posto", "posto_barca"),
        ("Cognome", "cognome"),
        ("Nome", "nome"),
        ("Codice Fiscale", "codice_fiscale"),
        ("Indirizzo", "indirizzo"),
        ("Telefono", "telefono"),
        ("Cellulare", "cellulare"),
        ("Email", "email"),
        ("Tipo barca", "tipo_barca"),
        ("Lunghezza (m)", "lunghezza"),
        ("Tipo sosta", "tipo_sosta"),
        ("Gg. sosta temp.", "giorni_sosta_temporanea"),
        ("Destinazione alaggio/varo", "destinazione_alaggio_varo"),
        ("Nome altra destinazione", "destinazione_altra_nome"),
        ("N° movimenti", "numero_movimenti"),
        ("Sosta €", "costo_sosta"),
        ("Movimentazione €", "costo_movimentazione"),
        ("Taccaggio €", "costo_taccaggio"),
        ("Copertura €", "costo_copertura"),
        ("Alaggio €", "costo_alaggio"),
        ("Varo €", "costo_varo"),
        ("Antivegetativa €", "costo_antivegetativa"),
        ("Magg. scafo sporco €", "costo_scafo_sporco"),
        ("Lavaggio inizio €", "costo_lavaggio_inizio"),
        ("Lavaggio fine €", "costo_lavaggio_fine"),
        ("Manutenzione motore €", "costo_manutenzione_motore"),
        ("Lavorazioni extra €", "__totale_extra__"),
        ("TOTALE €", "__totale__"),
        ("Pagato", "__pagato__"),
        ("Data pagamento", "data_pagamento"),
        ("Scad. antivegetativa", "scadenza_antivegetativa"),
        ("Scad. manutenzione", "scadenza_manutenzione"),
        ("Note lavori", "note_lavori"),
    ]

    COST_KEYS = ("costo_sosta","costo_movimentazione","costo_taccaggio","costo_copertura",
                 "costo_alaggio","costo_varo","costo_antivegetativa","costo_scafo_sporco",
                 "costo_lavaggio_inizio","costo_lavaggio_fine","costo_manutenzione_motore")

    def row_for(d: dict):
        tot_extra = round(sum(float((it or {}).get("prezzo") or 0) for it in (d.get("lavorazioni_extra") or [])), 2)
        totale = round(sum(float(d.get(k) or 0) for k in COST_KEYS) + tot_extra, 2)
        sosta_map = {"dentro": "Al coperto", "fuori": "Su piazzale", "fuori_sede": "Fuori sede", "temporanea": "Temporanea"}
        dest_map = {"marina_di_campo": "Marina di Campo", "altra": "Altra"}
        out = {}
        for label, key in COLS:
            if key == "__totale__":
                out[label] = totale
            elif key == "__totale_extra__":
                out[label] = tot_extra
            elif key == "__pagato__":
                out[label] = "Sì" if d.get("pagato") else "No"
            elif key == "tipo_sosta":
                out[label] = sosta_map.get(d.get("tipo_sosta"), d.get("tipo_sosta") or "")
            elif key == "destinazione_alaggio_varo":
                out[label] = dest_map.get(d.get("destinazione_alaggio_varo"), "")
            else:
                out[label] = d.get(key, "")
        return out

    rows = [row_for(d) for d in docs]
    df = pd.DataFrame(rows, columns=[c[0] for c in COLS]) if rows else pd.DataFrame(columns=[c[0] for c in COLS])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet_name = f"Clienti {anno}" if anno else "Clienti"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color="17324D", end_color="17324D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        totali_fill = PatternFill(start_color="B0562E", end_color="B0562E", fill_type="solid")
        totali_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(border_style="thin", color="D9D9D9")

        for col_idx, (label, _) in enumerate(COLS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        currency_labels = {label for label, _ in COLS if label.endswith("€")}
        for row_idx in range(2, len(rows) + 2):
            for col_idx, (label, _) in enumerate(COLS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if label in currency_labels:
                    cell.number_format = '#,##0.00 "€"'
                    cell.alignment = Alignment(horizontal="right")
                elif label == "Lunghezza (m)":
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right")
                if label == "TOTALE €":
                    cell.font = Font(bold=True)

        if rows:
            tot_row = len(rows) + 2
            ws.cell(row=tot_row, column=1, value="TOTALI").font = totali_font
            ws.cell(row=tot_row, column=1).fill = totali_fill
            ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="right")
            ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=15)
            for col_idx, (label, _) in enumerate(COLS, start=1):
                if label in currency_labels:
                    col_letter = get_column_letter(col_idx)
                    ws.cell(row=tot_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{tot_row-1})")
                    ws.cell(row=tot_row, column=col_idx).number_format = '#,##0.00 "€"'
                    ws.cell(row=tot_row, column=col_idx).font = totali_font
                    ws.cell(row=tot_row, column=col_idx).fill = totali_fill
                    ws.cell(row=tot_row, column=col_idx).alignment = Alignment(horizontal="right")

        widths = {
            "Anno": 8, "Posto": 8, "Cognome": 16, "Nome": 14, "Codice Fiscale": 20,
            "Indirizzo": 28, "Telefono": 14, "Cellulare": 14, "Email": 24,
            "Tipo barca": 18, "Lunghezza (m)": 12, "Tipo sosta": 14,
            "Gg. sosta temp.": 10, "Destinazione alaggio/varo": 22, "Nome altra destinazione": 22,
            "Pagato": 8, "Data pagamento": 14, "Scad. antivegetativa": 16, "Scad. manutenzione": 16,
            "Note lavori": 40,
        }
        for col_idx, (label, _) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(label, 15 if label.endswith("€") else 14)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    buf.seek(0)
    fname = f"clienti_cantiere_{anno}.xlsx" if anno else "clienti_cantiere.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/export/preventivi.zip")
async def export_tutti_pdf():
    """Esporta un archivio ZIP con un PDF preventivo per ogni cliente."""
    clienti_docs = await db.clienti.find({}, {"_id": 0}).to_list(10000)
    if not clienti_docs:
        raise HTTPException(404, "Nessun cliente da esportare")
    cantiere_doc = await db.cantiere.find_one({"id": "default"}, {"_id": 0}) or {}
    t_current = await get_tariffe_doc()

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for c in clienti_docs:
            lavori_docs = await db.lavori.find({"cliente_id": c["id"]}, {"_id": 0}).sort("data", -1).to_list(500)
            pdf_bytes = _build_preventivo_pdf(c, lavori_docs, cantiere_doc, t_current)
            posto = f"{int(c['posto_barca']):03d}_" if c.get("posto_barca") else ""
            filename = f"{posto}{(c.get('cognome') or 'cliente').lower()}_{(c.get('nome') or '').lower()}.pdf"
            filename = "".join(ch for ch in filename if ch.isalnum() or ch in "._-")
            zf.writestr(filename, pdf_bytes)
    zip_buf.seek(0)
    zip_filename = f"preventivi_cantiere_{datetime.now().strftime('%Y%m%d')}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={zip_filename}"}
    )
